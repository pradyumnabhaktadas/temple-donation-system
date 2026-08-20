import csv
import io
import os
import re
import secrets
import datetime
import threading
from collections import defaultdict
from functools import wraps

from flask import (
    Blueprint, render_template, request, redirect, url_for, flash,
    current_app, Response, abort,
)
from flask_login import login_user, logout_user, login_required, current_user
from sqlalchemy import func, extract

from extensions import db
from models import (
    Camp,
    Donor, Campaign, Donation, AdminUser, ReceiptCounter, BaceProperty, Festival, SevaType,
    LiveToGivePurpose, Preacher, AdminActivityLog, DONOR_TYPES, DONOR_TYPE_LABELS, DONATION_FREQUENCIES,
    DONATION_FREQUENCY_LABELS,
)
from utils import csv_safe_row, get_financial_year, is_valid_pan, is_valid_phone, normalize_phone, now_ist, to_ist
from pdf_utils import generate_receipt_pdf
from public import (
    find_or_create_donor, _org_cfg, high_value_pan_address_error, _finalize_success,
    _send_receipt_notifications_background,
)
from backup_utils import build_backup_zip, run_backup, restore_backup_zip

bp = Blueprint("admin", __name__, url_prefix="/admin")

ADMIN_ROLES = ["staff", "manager", "admin"]
ADMIN_ROLE_LABELS = {
    "staff": "Staff -- day-to-day work (log donations, view donors/reports)",
    "manager": "Manager -- staff permissions, same restrictions as staff for now",
    "admin": "Admin -- full access, including managing campaigns and other accounts",
}


def _generate_temp_password():
    """A random 12-character password for a newly created or reset admin
    account -- shown once in the results banner (there's no email on file
    for AdminUser to send it to), with must_change_password forced on so
    it can never be the account's permanent password."""
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghjkmnpqrstuvwxyz23456789"  # no 0/O/1/I/l
    return "".join(secrets.choice(alphabet) for _ in range(12))


def log_activity(action, target_type=None, target_id=None, details=None):
    """Records one row in the admin activity log (see AdminActivityLog in
    models.py) -- call this alongside the db.session change it's
    describing, before the commit, so the log entry lands in the same
    transaction as the change itself. Deliberately swallows its own
    errors: a logging failure should never break the actual admin action
    it's describing (e.g. a donor edit must still succeed even if, for
    some reason, writing the log row fails)."""
    try:
        db.session.add(AdminActivityLog(
            admin_username=current_user.username if current_user.is_authenticated else "system",
            action=action, target_type=target_type, target_id=target_id, details=details,
        ))
    except Exception:
        current_app.logger.exception("Failed to write admin activity log entry for action=%s", action)


def admin_role_required(view):
    """Restricts a route to the 'admin' role. 'staff' and 'manager' accounts
    can still do day-to-day work (log donations, view donors/reports) but
    can't create/edit campaigns, merge donor records, or manage other
    admin accounts -- those affect the whole organisation's data."""
    @wraps(view)
    def wrapped(*args, **kwargs):
        if current_user.role != "admin":
            flash("That action requires an administrator account.")
            return redirect(url_for("admin.dashboard"))
        return view(*args, **kwargs)
    return wrapped


@bp.before_request
def enforce_password_change():
    # Don't let a first-login admin wander off to other admin pages until
    # they've set a real password.
    if (
        current_user.is_authenticated
        and getattr(current_user, "must_change_password", False)
        and request.endpoint not in ("admin.change_password", "admin.logout")
    ):
        flash("Please set a new password before continuing.")
        return redirect(url_for("admin.change_password"))


@bp.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        user = AdminUser.query.filter_by(username=username).first()

        if user and user.is_locked():
            minutes_left = max(1, int((user.locked_until - datetime.datetime.utcnow()).total_seconds() // 60) + 1)
            flash(f"Too many failed attempts. Try again in about {minutes_left} minute(s).")
            return render_template("admin/login.html")

        if user and user.check_password(password):
            user.register_successful_login()
            db.session.commit()
            login_user(user)
            if user.must_change_password:
                flash("Please set a new password before continuing.")
                return redirect(url_for("admin.change_password"))
            return redirect(url_for("admin.dashboard"))

        if user:
            max_attempts = current_app.config["LOGIN_MAX_ATTEMPTS"]
            lockout_minutes = current_app.config["LOGIN_LOCKOUT_MINUTES"]
            user.register_failed_attempt(max_attempts, lockout_minutes)
            db.session.commit()
        flash("Invalid username or password")
    return render_template("admin/login.html")


@bp.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("admin.login"))


@bp.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not current_user.check_password(current_password):
            flash("Current password is incorrect.")
        elif len(new_password) < 8:
            flash("New password must be at least 8 characters.")
        elif new_password != confirm_password:
            flash("New password and confirmation don't match.")
        else:
            current_user.set_password(new_password)
            current_user.must_change_password = False
            db.session.commit()
            flash("Password updated.")
            return redirect(url_for("admin.dashboard"))

    return render_template("admin/change_password.html", force=current_user.must_change_password)


@bp.route("/settings/users")
@login_required
@admin_role_required
def manage_users():
    """Admin-only account management -- add staff/manager/admin logins,
    reset a forgotten password, unlock a rate-limited account, change a
    role, or remove a former staff member's access. Previously the only
    way to do any of this was direct database/shell access (e.g. re-running
    seed.py), which doesn't scale past one admin."""
    users = AdminUser.query.order_by(AdminUser.role.desc(), AdminUser.username).all()
    return render_template(
        "admin/manage_users.html", users=users,
        admin_roles=ADMIN_ROLES, admin_role_labels=ADMIN_ROLE_LABELS,
    )


@bp.route("/settings/users/add", methods=["POST"])
@login_required
@admin_role_required
def add_user():
    username = request.form.get("username", "").strip()
    role = request.form.get("role", "staff")

    if not username:
        flash("Username is required.")
        return redirect(url_for("admin.manage_users"))
    if role not in ADMIN_ROLES:
        role = "staff"
    if AdminUser.query.filter_by(username=username).first():
        flash(f"Username '{username}' is already taken.")
        return redirect(url_for("admin.manage_users"))

    temp_password = _generate_temp_password()
    user = AdminUser(username=username, role=role, must_change_password=True)
    user.set_password(temp_password)
    db.session.add(user)
    db.session.flush()
    log_activity("admin_user_add", target_type="admin_user", target_id=user.id, details=f"Created account '{username}' ({role})")
    db.session.commit()

    flash(
        f"Account '{username}' created ({role}). Temporary password: {temp_password} -- "
        "share this with them directly (it won't be shown again); they'll be required to "
        "set their own password on first login."
    )
    return redirect(url_for("admin.manage_users"))


@bp.route("/settings/users/<int:user_id>/reset-password", methods=["POST"])
@login_required
@admin_role_required
def reset_user_password(user_id):
    user = AdminUser.query.get_or_404(user_id)
    temp_password = _generate_temp_password()
    user.set_password(temp_password)
    user.must_change_password = True
    user.failed_attempts = 0
    user.locked_until = None
    log_activity("admin_user_reset_password", target_type="admin_user", target_id=user.id, details=f"Reset password for '{user.username}'")
    db.session.commit()

    flash(
        f"Password reset for '{user.username}'. Temporary password: {temp_password} -- "
        "share this with them directly (it won't be shown again); they'll be required to "
        "set their own password on next login."
    )
    return redirect(url_for("admin.manage_users"))


@bp.route("/settings/users/<int:user_id>/unlock", methods=["POST"])
@login_required
@admin_role_required
def unlock_user(user_id):
    user = AdminUser.query.get_or_404(user_id)
    user.failed_attempts = 0
    user.locked_until = None
    log_activity("admin_user_unlock", target_type="admin_user", target_id=user.id, details=f"Unlocked '{user.username}'")
    db.session.commit()
    flash(f"'{user.username}' has been unlocked.")
    return redirect(url_for("admin.manage_users"))


@bp.route("/settings/users/<int:user_id>/role", methods=["POST"])
@login_required
@admin_role_required
def change_user_role(user_id):
    user = AdminUser.query.get_or_404(user_id)
    new_role = request.form.get("role", "")

    if new_role not in ADMIN_ROLES:
        flash("Invalid role.")
        return redirect(url_for("admin.manage_users"))

    if user.id == current_user.id and new_role != "admin":
        # An admin demoting themselves is how you accidentally lock
        # yourself out of Settings/user management with nobody left able
        # to undo it (if they're also the last admin) -- block outright
        # rather than trying to detect "are you the last one" here too.
        flash("You can't change your own role away from admin. Have another admin do this for you.")
        return redirect(url_for("admin.manage_users"))

    old_role = user.role
    user.role = new_role
    log_activity("admin_user_role_change", target_type="admin_user", target_id=user.id, details=f"'{user.username}' role changed: {old_role} -> {new_role}")
    db.session.commit()
    flash(f"'{user.username}' is now {new_role}.")
    return redirect(url_for("admin.manage_users"))


@bp.route("/settings/users/<int:user_id>/delete", methods=["POST"])
@login_required
@admin_role_required
def delete_user(user_id):
    user = AdminUser.query.get_or_404(user_id)

    if user.id == current_user.id:
        flash("You can't delete your own account while logged in as it.")
        return redirect(url_for("admin.manage_users"))

    if user.role == "admin" and AdminUser.query.filter_by(role="admin").count() <= 1:
        flash("Can't delete the last remaining admin account -- promote another account to admin first.")
        return redirect(url_for("admin.manage_users"))

    # AdminUser isn't referenced by a foreign key anywhere (donation.
    # recorded_by is just a text snapshot like "manual entry (username)"),
    # so deleting the account doesn't touch any donation/donor history.
    username, deleted_id = user.username, user.id
    db.session.delete(user)
    log_activity("admin_user_delete", target_type="admin_user", target_id=deleted_id, details=f"Deleted account '{username}'")
    db.session.commit()
    flash(f"Account '{username}' has been removed.")
    return redirect(url_for("admin.manage_users"))


ACTIVITY_LOG_PER_PAGE = 50


@bp.route("/settings/activity-log")
@login_required
@admin_role_required
def activity_log():
    """Admin-only page listing every row written by log_activity() above --
    donor edits/merges, donation cancel/restore, campaign CRUD, and admin
    user management actions, newest first. Filterable by which admin did
    it and what kind of action, so "what did X do last week" or "who
    cancelled this donation" are both answerable without a database
    console."""
    query = AdminActivityLog.query

    admin_username = request.args.get("admin_username", "")
    if admin_username:
        query = query.filter_by(admin_username=admin_username)

    action = request.args.get("action", "")
    if action:
        query = query.filter_by(action=action)

    page = request.args.get("page", 1, type=int)
    pagination = db.paginate(
        query.order_by(AdminActivityLog.created_at.desc()), page=page, per_page=ACTIVITY_LOG_PER_PAGE, error_out=False
    )

    # Options for the filter dropdowns -- every admin username/action that's
    # actually appeared in the log so far, rather than a hardcoded list that
    # could drift out of sync with what's really been logged.
    known_usernames = [r[0] for r in db.session.query(AdminActivityLog.admin_username).distinct().order_by(AdminActivityLog.admin_username).all()]
    known_actions = [r[0] for r in db.session.query(AdminActivityLog.action).distinct().order_by(AdminActivityLog.action).all()]

    return render_template(
        "admin/activity_log.html", pagination=pagination, entries=pagination.items,
        admin_username=admin_username, action=action,
        known_usernames=known_usernames, known_actions=known_actions,
    )


@bp.route("/")
@bp.route("/dashboard")
@login_required
def dashboard():
    # now_ist(), not datetime.date.today() -- the server's own clock is UTC
    # (e.g. on Render), so "today" would read as the wrong calendar date for
    # the ~5.5 hours a day (roughly 12:00 AM-5:30 AM IST) where UTC hasn't
    # rolled over to the same day yet.
    today = now_ist().date()
    start_of_month = today.replace(day=1)
    fy = get_financial_year(today)
    fy_start_year = int(fy.split("-")[0])
    fy_start = datetime.date(fy_start_year, 4, 1)

    def total_since(d):
        dt = datetime.datetime.combine(d, datetime.time.min)
        return db.session.query(func.coalesce(func.sum(Donation.amount), 0)).filter(
            Donation.status == "success", Donation.donation_date >= dt
        ).scalar()

    today_start = datetime.datetime.combine(today, datetime.time.min)
    today_end = today_start + datetime.timedelta(days=1)
    today_total = db.session.query(func.coalesce(func.sum(Donation.amount), 0)).filter(
        Donation.status == "success",
        Donation.donation_date >= today_start,
        Donation.donation_date < today_end,
    ).scalar()
    month_total = total_since(start_of_month)
    year_total = total_since(fy_start)

    campaign_totals = (
        db.session.query(Campaign.name, func.coalesce(func.sum(Donation.amount), 0))
        .outerjoin(Donation, (Donation.campaign_id == Campaign.id) & (Donation.status == "success"))
        .group_by(Campaign.id)
        .order_by(func.coalesce(func.sum(Donation.amount), 0).desc())
        .all()
    )

    mode_totals = (
        db.session.query(Donation.payment_mode, func.coalesce(func.sum(Donation.amount), 0))
        .filter(Donation.status == "success")
        .group_by(Donation.payment_mode)
        .all()
    )

    # last 6 months trend
    def months_ago(base, i):
        """Calendar-correct 'i months before base', avoiding drift from
        fixed 30-day subtraction across months of different lengths."""
        total_months = base.year * 12 + (base.month - 1) - i
        y, m = divmod(total_months, 12)
        return datetime.date(y, m + 1, 1)

    monthly = []
    for i in range(5, -1, -1):
        month_date = months_ago(start_of_month, i)
        y, m = month_date.year, month_date.month
        total = db.session.query(func.coalesce(func.sum(Donation.amount), 0)).filter(
            Donation.status == "success",
            extract("year", Donation.donation_date) == y,
            extract("month", Donation.donation_date) == m,
        ).scalar()
        monthly.append({"label": month_date.strftime("%b %Y"), "total": float(total)})

    donor_count = Donor.query.count()
    donation_count = Donation.query.filter_by(status="success").count()

    # Form 10BD reminder: the statutory deadline for the annual "statement
    # of donations" filing (covering the FY that just ended on 31 March) is
    # 31 May. Surface a reminder banner during that April-May filing window.
    form_10bd_reminder = None
    if today.month in (4, 5):
        deadline = datetime.date(today.year, 5, 31)
        filing_fy = f"{today.year - 1}-{str(today.year)[-2:]}"
        days_left = (deadline - today).days
        form_10bd_reminder = {"filing_fy": filing_fy, "days_left": days_left, "overdue": days_left < 0}

    # Birthday reminder banner -- today's only; the full calendar with
    # upcoming birthdays lives on its own page (Admin -> Donors -> Birthdays,
    # see the birthdays() route below). Reuses the same Feb-29/year-
    # wraparound-safe date math as Donor Insights and Analytics
    # (_next_occurrence, defined further down in this file).
    birthdays_today = []
    for donor in Donor.query.filter(Donor.dob.isnot(None)).all():
        days_until, _occurrence_date = _next_occurrence(donor.dob, today)
        if days_until == 0:
            birthdays_today.append({"donor": donor})

    # Failed/abandoned online donations -- a Donation row is created with
    # status="pending" the instant checkout starts (see
    # public.create_order), before the payment actually confirms, so a
    # donor who closes the Razorpay popup or whose payment fails leaves a
    # real row behind that's easy to miss (the Donations Log defaults to
    # status=success, see _apply_donations_filters above). Surface the
    # last 7 days of these here so staff can proactively follow up --
    # "still pending" often just means "abandoned the checkout partway",
    # worth a call/WhatsApp nudge; "failed" means the payment itself
    # didn't go through. Excludes anything from the last 30 minutes so a
    # donor who is mid-checkout right now doesn't show up as if something
    # were already wrong.
    stale_cutoff = datetime.datetime.utcnow() - datetime.timedelta(minutes=30)
    lookback = datetime.datetime.utcnow() - datetime.timedelta(days=7)
    failed_abandoned_donations = (
        Donation.query.filter(
            Donation.status.in_(["pending", "failed"]),
            Donation.donation_date >= lookback,
            Donation.donation_date <= stale_cutoff,
        )
        .order_by(Donation.donation_date.desc())
        .limit(15)
        .all()
    )
    failed_abandoned_count = (
        Donation.query.filter(
            Donation.status.in_(["pending", "failed"]),
            Donation.donation_date >= lookback,
            Donation.donation_date <= stale_cutoff,
        ).count()
    )

    # Disputed/charged-back donations -- populated only by the webhook's
    # payment.dispute.* events (see public._handle_payment_dispute). A
    # dispute against a donation we've already captured (and possibly
    # issued an 80G receipt for) needs a human to look at it, so surface
    # every one that isn't yet in a resolved state ("won"/"lost"/"closed")
    # -- those stay visible in the Donations Log / donor detail page but
    # don't need to keep occupying a Dashboard alert once Razorpay's own
    # process has concluded.
    disputed_donations = (
        Donation.query.filter(
            Donation.razorpay_dispute_id.isnot(None),
            ~Donation.razorpay_dispute_status.in_(["won", "lost", "closed"]),
        )
        .order_by(Donation.disputed_at.desc())
        .limit(15)
        .all()
    )

    return render_template(
        "admin/dashboard.html",
        today_total=float(today_total),
        month_total=float(month_total),
        year_total=float(year_total),
        campaign_totals=[(n, float(t)) for n, t in campaign_totals],
        mode_totals=[(m, float(t)) for m, t in mode_totals],
        monthly=monthly,
        donor_count=donor_count,
        donation_count=donation_count,
        fy=fy,
        form_10bd_reminder=form_10bd_reminder,
        birthdays_today=birthdays_today,
        failed_abandoned_donations=failed_abandoned_donations,
        failed_abandoned_count=failed_abandoned_count,
        disputed_donations=disputed_donations,
        today=today,
    )


def _months_ago(base, i):
    """Calendar-correct 'i months before base', avoiding drift from fixed
    30-day subtraction across months of different lengths."""
    total_months = base.year * 12 + (base.month - 1) - i
    y, m = divmod(total_months, 12)
    return datetime.date(y, m + 1, 1)


def _apply_donor_population_filters(query, donor_type, preacher_id, frequency):
    """Shared donor_type/connected_preacher/donation_frequency filtering,
    used by both the Donors list and Analytics -- `query` must have Donor
    columns in scope (either Donor.query directly, or a query that's
    joined Donor in). "none" for preacher_id is a sentinel meaning "no
    preacher assigned" (IS NULL), distinct from a blank/absent param
    ("any preacher")."""
    if donor_type:
        query = query.filter(Donor.donor_type == donor_type)
    if preacher_id == "none":
        query = query.filter(Donor.connected_preacher_id.is_(None))
    elif preacher_id:
        try:
            query = query.filter(Donor.connected_preacher_id == int(preacher_id))
        except ValueError:
            pass
    if frequency:
        query = query.filter(Donor.donation_frequency == frequency)
    return query


# Retention thresholds (days since last successful donation) -- a donor's
# frequency label (set manually by staff, see Donor.donation_frequency)
# determines when a gap starts looking like a missed gift vs. business as
# usual. These are reasonable defaults, not a spec from the temple --
# tune freely if real-world follow-up patterns suggest otherwise.
_RETENTION_THRESHOLDS = {
    "monthly": (35, 60),      # active <=35 days, at-risk 36-60, inactive >60
    "quarterly": (100, 150),
    "yearly": (380, 450),
    None: (180, 365),         # occasional / one_time / not set
}
for _f in ("occasional", "one_time"):
    _RETENTION_THRESHOLDS[_f] = _RETENTION_THRESHOLDS[None]


def _donor_activity_status(frequency, last_donation_date, today):
    """One of "never", "active", "at_risk", "inactive" for a donor, based
    on how long it's been since their last successful donation relative
    to how often they're expected to give."""
    if last_donation_date is None:
        return "never"
    days_since = (today - last_donation_date).days
    at_risk_after, inactive_after = _RETENTION_THRESHOLDS.get(frequency, _RETENTION_THRESHOLDS[None])
    if days_since <= at_risk_after:
        return "active"
    elif days_since <= inactive_after:
        return "at_risk"
    return "inactive"


def _resolve_analytics_date_range(preset, custom_from_raw, custom_to_raw, today):
    """Resolves the Analytics page's date-range filter to a [start, end)
    pair of dates. "this_quarter"/"this_year" are to-date (like the
    Dashboard's FY total) rather than the full period including days that
    haven't happened yet; "last_month" is the one preset that's a fully
    closed period."""
    if preset == "last_month":
        end_of_last_month = today.replace(day=1) - datetime.timedelta(days=1)
        start = end_of_last_month.replace(day=1)
        end = today.replace(day=1)
        label = end_of_last_month.strftime("%B %Y")
    elif preset == "this_quarter":
        q_start_month = ((today.month - 1) // 3) * 3 + 1
        start = today.replace(month=q_start_month, day=1)
        end = today + datetime.timedelta(days=1)
        label = f"This Quarter (from {start.strftime('%d-%b')})"
    elif preset == "this_year":
        fy = get_financial_year(today)
        start = datetime.date(int(fy.split("-")[0]), 4, 1)
        end = today + datetime.timedelta(days=1)
        label = f"FY {fy} to date"
    elif preset == "custom":
        try:
            start = datetime.datetime.strptime(custom_from_raw, "%Y-%m-%d").date() if custom_from_raw else today - datetime.timedelta(days=30)
        except ValueError:
            start = today - datetime.timedelta(days=30)
        try:
            custom_to = datetime.datetime.strptime(custom_to_raw, "%Y-%m-%d").date() if custom_to_raw else today
        except ValueError:
            custom_to = today
        end = custom_to + datetime.timedelta(days=1)
        label = f"{start.strftime('%d-%b-%Y')} to {custom_to.strftime('%d-%b-%Y')}"
    else:
        preset = "this_month"
        start = today.replace(day=1)
        end = today + datetime.timedelta(days=1)
        label = today.strftime("%B %Y")
    return preset, start, end, label


@bp.route("/analytics")
@login_required
def analytics():
    """Analytics built around the four things that actually matter for
    running an IYF/Live To Give donor program: how much is coming in
    (Donations), how the donor base is growing (Donors), how relationships
    are distributed (Preachers), and who needs attention right now
    (Follow-up) -- not just a running total.

    Implementation note: rather than issuing a separate SQL query per
    section, this pulls the (small, single-temple-scale) filtered donor
    list and their full donation history once each, then does every
    breakdown/aggregate in Python. That keeps ~12 report sections easy to
    verify and keeps the SQL portable (no dialect-specific DISTINCT ON /
    correlated-subquery tricks needed across SQLite dev / Postgres prod).
    """
    # now_ist(), not datetime.date.today() -- the server's own clock is UTC
    # (e.g. on Render), so "today" would read as the wrong calendar date for
    # the ~5.5 hours a day (roughly 12:00 AM-5:30 AM IST) where UTC hasn't
    # rolled over to the same day yet.
    today = now_ist().date()

    # -- Global filters (apply across every section) --
    date_preset = request.args.get("date_preset", "this_month")
    custom_from_raw = request.args.get("date_from", "")
    custom_to_raw = request.args.get("date_to", "")
    date_preset, period_start, period_end, period_label = _resolve_analytics_date_range(
        date_preset, custom_from_raw, custom_to_raw, today
    )

    donor_type_filter = request.args.get("donor_type", "")
    preacher_filter = request.args.get("preacher_id", "")
    frequency_filter = request.args.get("donation_frequency", "")

    donors_all = _apply_donor_population_filters(
        Donor.query, donor_type_filter, preacher_filter, frequency_filter
    ).all()
    population_donor_ids = [d.id for d in donors_all]
    donor_by_id = {d.id: d for d in donors_all}
    all_preachers = Preacher.query.order_by(Preacher.name).all()
    preacher_name_by_id = {p.id: p.name for p in all_preachers}

    trend_granularity = request.args.get("trend", "monthly")
    if trend_granularity not in ("monthly", "quarterly", "yearly"):
        trend_granularity = "monthly"
    top_scope = request.args.get("top_scope", "lifetime")
    if top_scope not in ("lifetime", "period"):
        top_scope = "lifetime"

    # Every successful donation by a population-filtered donor, pulled
    # once as (donor_id, amount, plain date) tuples -- everything below
    # slices this same list rather than re-querying.
    all_pop_donations = []
    if population_donor_ids:
        for did, amt, dt in (
            db.session.query(Donation.donor_id, Donation.amount, Donation.donation_date)
            .filter(Donation.status == "success", Donation.donor_id.in_(population_donor_ids))
            .all()
        ):
            d = dt.date() if hasattr(dt, "date") else dt
            all_pop_donations.append((did, float(amt), d))

    period_donation_rows = [(did, amt) for did, amt, d in all_pop_donations if period_start <= d < period_end]
    total_donations_amount = sum(amt for _, amt in period_donation_rows)
    active_donor_ids = {did for did, _ in period_donation_rows}
    donation_count_in_period = len(period_donation_rows)
    avg_donation_in_period = (total_donations_amount / donation_count_in_period) if donation_count_in_period else 0

    this_month_start = today.replace(day=1)
    this_month_total = sum(amt for _, amt, d in all_pop_donations if d >= this_month_start)

    first_donation_date, last_donation_date = {}, {}
    for did, amt, d in all_pop_donations:
        if did not in first_donation_date or d < first_donation_date[did]:
            first_donation_date[did] = d
        if did not in last_donation_date or d > last_donation_date[did]:
            last_donation_date[did] = d

    new_donors_count = sum(1 for d in first_donation_date.values() if period_start <= d < period_end)

    def _created_before(donor, cutoff_date):
        created = donor.created_at.date() if donor.created_at else datetime.date.min
        return created < cutoff_date

    # ================= 1. KPI CARDS =================
    kpis = {
        "total_donors": sum(1 for d in donors_all if _created_before(d, period_end)),
        "active_donors": len(active_donor_ids),
        "total_donations": total_donations_amount,
        "this_month": this_month_total,
        "monthly_recurring_donors": sum(
            1 for d in donors_all if d.donation_frequency == "monthly" and _created_before(d, period_end)
        ),
        "new_donors": new_donors_count,
        "avg_donation": avg_donation_in_period,
        "preacher_count": len({d.connected_preacher_id for d in donors_all if d.connected_preacher_id}),
    }

    # ================= 2. DONATION TREND =================
    trend_buckets = []
    if trend_granularity == "monthly":
        for i in range(11, -1, -1):
            bucket_start = _months_ago(today.replace(day=1), i)
            bucket_end = _months_ago(today.replace(day=1), i - 1)
            trend_buckets.append((bucket_start.strftime("%b %Y"), bucket_start, bucket_end))
    elif trend_granularity == "quarterly":
        this_q_start_month = ((today.month - 1) // 3) * 3 + 1
        this_q_start = today.replace(month=this_q_start_month, day=1)
        for i in range(7, -1, -1):
            bucket_start = _months_ago(this_q_start, i * 3)
            bucket_end = _months_ago(this_q_start, (i - 1) * 3)
            q_num = (bucket_start.month - 1) // 3 + 1
            trend_buckets.append((f"Q{q_num} {bucket_start.year}", bucket_start, bucket_end))
    else:  # yearly (financial years)
        current_fy_start_year = int(get_financial_year(today).split("-")[0])
        for offset in range(4, -1, -1):
            fy_start_year = current_fy_start_year - offset
            bucket_start = datetime.date(fy_start_year, 4, 1)
            bucket_end = datetime.date(fy_start_year + 1, 4, 1)
            trend_buckets.append((f"FY {fy_start_year}-{str(fy_start_year + 1)[-2:]}", bucket_start, bucket_end))

    donation_trend = []
    for label, bstart, bend in trend_buckets:
        bucket_rows = [amt for _, amt, d in all_pop_donations if bstart <= d < bend]
        donation_trend.append({"label": label, "total": sum(bucket_rows), "count": len(bucket_rows)})

    # ================= 3. DONOR TYPE ANALYTICS =================
    donor_amount_by_id = {}
    for did, amt in period_donation_rows:
        donor_amount_by_id[did] = donor_amount_by_id.get(did, 0) + amt

    type_totals = {t: {"donors": 0, "amount": 0.0} for t in DONOR_TYPES}
    for d in donors_all:
        if d.donor_type in type_totals:
            type_totals[d.donor_type]["donors"] += 1
            type_totals[d.donor_type]["amount"] += donor_amount_by_id.get(d.id, 0)
    grand_total_amount = sum(v["amount"] for v in type_totals.values()) or 1.0
    donor_type_breakdown = [
        {
            "type": t, "label": DONOR_TYPE_LABELS[t],
            "donors": type_totals[t]["donors"], "amount": type_totals[t]["amount"],
            "pct": round(type_totals[t]["amount"] / grand_total_amount * 100, 1),
        }
        for t in DONOR_TYPES
    ]

    # ================= 4. DONATION FREQUENCY BREAKDOWN =================
    freq_totals = {f: {"donors": 0, "amount": 0.0} for f in DONATION_FREQUENCIES}
    for d in donors_all:
        if d.donation_frequency in freq_totals:
            freq_totals[d.donation_frequency]["donors"] += 1
            freq_totals[d.donation_frequency]["amount"] += donor_amount_by_id.get(d.id, 0)
    frequency_breakdown = [
        {
            "freq": f, "label": DONATION_FREQUENCY_LABELS[f],
            "donors": freq_totals[f]["donors"], "amount": freq_totals[f]["amount"],
        }
        for f in DONATION_FREQUENCIES
    ]

    # ================= 5. PREACHER-WISE PERFORMANCE =================
    donors_by_preacher = {}
    for d in donors_all:
        donors_by_preacher.setdefault(d.connected_preacher_id, []).append(d)

    preacher_performance = []
    for p in all_preachers:
        p_donors = donors_by_preacher.get(p.id, [])
        if not p_donors:
            continue
        p_ids = {d.id for d in p_donors}
        lifetime_rows = [amt for did, amt, d in all_pop_donations if did in p_ids]
        lifetime_total = sum(lifetime_rows)
        lifetime_count = len(lifetime_rows)
        preacher_performance.append({
            "id": p.id, "name": p.name, "donors": len(p_donors),
            "active_donors": sum(1 for did in p_ids if did in active_donor_ids),
            "period_total": sum(amt for did, amt in period_donation_rows if did in p_ids),
            "lifetime_total": lifetime_total,
            "monthly_donors": sum(1 for d in p_donors if d.donation_frequency == "monthly"),
            "avg_donation": (lifetime_total / lifetime_count) if lifetime_count else 0,
        })
    preacher_performance.sort(key=lambda r: r["lifetime_total"], reverse=True)

    # ================= 6. DONOR GROWTH =================
    fy_start_date = datetime.date(int(get_financial_year(today).split("-")[0]), 4, 1)
    q_start_month = ((today.month - 1) // 3) * 3 + 1
    q_start_date = today.replace(month=q_start_month, day=1)
    month_start_date = today.replace(day=1)

    new_this_month = sum(1 for d in first_donation_date.values() if d >= month_start_date)
    new_this_quarter = sum(1 for d in first_donation_date.values() if d >= q_start_date)
    new_this_year = sum(1 for d in first_donation_date.values() if d >= fy_start_date)

    growth_trend = []
    for i in range(11, -1, -1):
        bucket_start = _months_ago(today.replace(day=1), i)
        bucket_end = _months_ago(today.replace(day=1), i - 1)
        count = sum(1 for d in first_donation_date.values() if bucket_start <= d < bucket_end)
        growth_trend.append({"label": bucket_start.strftime("%b %Y"), "new": count})

    prev_period_len = (period_end - period_start).days
    prev_period_start = period_start - datetime.timedelta(days=prev_period_len)
    prev_period_new = sum(1 for d in first_donation_date.values() if prev_period_start <= d < period_start)
    donor_growth_pct = None
    if prev_period_new > 0:
        donor_growth_pct = round((new_donors_count - prev_period_new) / prev_period_new * 100, 1)

    # ================= 7. DONOR RETENTION =================
    retention_counts = {"active": 0, "at_risk": 0, "inactive": 0, "never": 0}
    at_risk_list = []
    for donor in donors_all:
        last_date = last_donation_date.get(donor.id)
        status = _donor_activity_status(donor.donation_frequency, last_date, today)
        retention_counts[status] += 1
        if status in ("at_risk", "inactive"):
            at_risk_list.append({
                "donor": donor, "status": status, "last_donation": last_date,
                "days_since": (today - last_date).days if last_date else None,
            })
    at_risk_list.sort(key=lambda r: r["days_since"] or 0, reverse=True)

    # ================= 8. BIRTHDAYS & ANNIVERSARIES =================
    donors_with_dates = [
        d for d in donors_all
        if d.dob or d.father_dob or d.mother_dob or d.wife_dob or d.marriage_anniversary
    ]
    upcoming_30 = []
    for donor in donors_with_dates:
        for label, date_value in [
            ("Donor's Birthday", donor.dob), ("Marriage Anniversary", donor.marriage_anniversary),
            ("Father's Birthday", donor.father_dob), ("Mother's Birthday", donor.mother_dob),
            ("Wife's Birthday", donor.wife_dob),
        ]:
            if not date_value:
                continue
            days_until, occurrence_date = _next_occurrence(date_value, today)
            if days_until is not None and days_until <= 30:
                upcoming_30.append({"donor": donor, "label": label, "days_until": days_until, "date": occurrence_date})
    upcoming_30.sort(key=lambda u: u["days_until"])
    upcoming_7 = [u for u in upcoming_30 if u["days_until"] <= 7]

    birthday_month_counts = {
        "Donor's Birthday": 0, "Father's Birthday": 0, "Mother's Birthday": 0,
        "Wife's Birthday": 0, "Marriage Anniversary": 0,
    }
    for u in upcoming_30:
        birthday_month_counts[u["label"]] += 1

    # ================= 9. GEOGRAPHIC ANALYTICS =================
    geo = {}
    for d in donors_all:
        state = (d.state or "").strip() or "Unknown"
        geo.setdefault(state, {"donors": 0, "amount": 0.0})
        geo[state]["donors"] += 1
    for did, amt in period_donation_rows:
        d = donor_by_id.get(did)
        state = ((d.state or "").strip() or "Unknown") if d else "Unknown"
        geo.setdefault(state, {"donors": 0, "amount": 0.0})
        geo[state]["amount"] += amt
    geography = sorted(
        [{"state": s, "donors": v["donors"], "amount": v["amount"]} for s, v in geo.items()],
        key=lambda r: r["amount"], reverse=True,
    )

    # ================= 10. DATA COMPLETENESS =================
    total_pop = len(donors_all) or 1
    with_pan = sum(1 for d in donors_all if d.pan)
    with_address = sum(1 for d in donors_all if d.address and d.city and d.state and d.pincode)
    with_dob = sum(1 for d in donors_all if d.dob)
    with_preacher = sum(1 for d in donors_all if d.connected_preacher_id)
    incomplete_count = 0
    for d in donors_all:
        missing = sum([
            not d.pan, not (d.address and d.city and d.state and d.pincode),
            not d.dob, not d.connected_preacher_id, not d.donor_type,
        ])
        if missing >= 2:
            incomplete_count += 1
    completeness = {
        "pan_pct": round(with_pan / total_pop * 100, 1),
        "address_pct": round(with_address / total_pop * 100, 1),
        "dob_pct": round(with_dob / total_pop * 100, 1),
        "preacher_pct": round(with_preacher / total_pop * 100, 1),
        "overall_pct": round((with_pan + with_address + with_dob + with_preacher) / (total_pop * 4) * 100, 1),
        "incomplete_count": incomplete_count,
    }

    # ================= 11. FOLLOW-UP DASHBOARD =================
    overdue_count = sum(
        1 for r in at_risk_list if r["donor"].donation_frequency in ("monthly", "quarterly", "yearly")
    )
    unassigned_count = sum(1 for d in donors_all if not d.connected_preacher_id)
    new_without_preacher = sum(
        1 for did, fdate in first_donation_date.items()
        if (today - fdate).days <= 30 and donor_by_id.get(did) and not donor_by_id[did].connected_preacher_id
    )
    lifetime_totals_by_donor = {}
    for did, amt, d in all_pop_donations:
        lifetime_totals_by_donor[did] = lifetime_totals_by_donor.get(did, 0) + amt
    sorted_totals = sorted(lifetime_totals_by_donor.values(), reverse=True)
    high_value_cutoff = sorted_totals[9] if len(sorted_totals) >= 10 else (sorted_totals[-1] if sorted_totals else 0)
    high_value_at_risk = [
        r for r in at_risk_list
        if high_value_cutoff > 0 and lifetime_totals_by_donor.get(r["donor"].id, 0) >= high_value_cutoff
    ][:10]

    followup = {
        "overdue_count": overdue_count,
        "unassigned_count": unassigned_count,
        "new_without_preacher": new_without_preacher,
        "inactive_count": retention_counts["inactive"],
        "birthdays_week_count": sum(1 for u in upcoming_7 if "Birthday" in u["label"]),
        "anniversaries_week_count": sum(1 for u in upcoming_7 if u["label"] == "Marriage Anniversary"),
        "incomplete_count": completeness["incomplete_count"],
        "high_value_at_risk": high_value_at_risk,
    }

    # ================= 12. TOP DONORS =================
    totals_source = period_donation_rows if top_scope == "period" else [(did, amt) for did, amt, d in all_pop_donations]
    totals_by_donor, counts_by_donor = {}, {}
    for did, amt in totals_source:
        totals_by_donor[did] = totals_by_donor.get(did, 0) + amt
        counts_by_donor[did] = counts_by_donor.get(did, 0) + 1
    top_ids = sorted(totals_by_donor, key=lambda did: totals_by_donor[did], reverse=True)[:10]
    top_donors = [
        {
            "id": did, "name": donor_by_id[did].full_name,
            "type": DONOR_TYPE_LABELS.get(donor_by_id[did].donor_type, "-"),
            "preacher": preacher_name_by_id.get(donor_by_id[did].connected_preacher_id, "-"),
            "total": totals_by_donor[did], "count": counts_by_donor[did],
        }
        for did in top_ids if did in donor_by_id
    ]

    return render_template(
        "admin/analytics.html",
        date_preset=date_preset, custom_from=custom_from_raw, custom_to=custom_to_raw, period_label=period_label,
        donor_type_filter=donor_type_filter, preacher_filter=preacher_filter, frequency_filter=frequency_filter,
        donor_types=DONOR_TYPES, donor_type_labels=DONOR_TYPE_LABELS,
        donation_frequencies=DONATION_FREQUENCIES, donation_frequency_labels=DONATION_FREQUENCY_LABELS,
        all_preachers=all_preachers,
        kpis=kpis,
        trend_granularity=trend_granularity, donation_trend=donation_trend,
        donor_type_breakdown=donor_type_breakdown,
        frequency_breakdown=frequency_breakdown,
        preacher_performance=preacher_performance,
        new_this_month=new_this_month, new_this_quarter=new_this_quarter, new_this_year=new_this_year,
        donor_growth_pct=donor_growth_pct, growth_trend=growth_trend,
        retention_counts=retention_counts, at_risk_list=at_risk_list[:25],
        upcoming_7=upcoming_7, upcoming_30=upcoming_30[:25], birthday_month_counts=birthday_month_counts,
        geography=geography,
        completeness=completeness,
        followup=followup,
        top_donors=top_donors, top_scope=top_scope,
    )


DONORS_PER_PAGE = 30
DONATIONS_PER_PAGE = 50


@bp.route("/donors")
@login_required
def donors():
    q = request.args.get("q", "").strip()
    donor_type = request.args.get("donor_type", "")
    # "none" is a sentinel meaning "no preacher assigned" (connected_preacher_id
    # IS NULL) -- distinct from a blank/absent param, which means "any".
    preacher_id = request.args.get("preacher_id", "")
    donation_frequency = request.args.get("donation_frequency", "")
    page = request.args.get("page", 1, type=int)
    query = Donor.query
    if q:
        like = f"%{q}%"
        query = query.filter(
            (Donor.full_name.ilike(like))
            | (Donor.phone.ilike(like))
            | (Donor.email.ilike(like))
            | (Donor.pan.ilike(like))
        )
    if donor_type:
        query = query.filter(Donor.donor_type == donor_type)
    if preacher_id == "none":
        query = query.filter(Donor.connected_preacher_id.is_(None))
    elif preacher_id:
        try:
            query = query.filter(Donor.connected_preacher_id == int(preacher_id))
        except ValueError:
            pass
    if donation_frequency:
        query = query.filter(Donor.donation_frequency == donation_frequency)
    pagination = db.paginate(
        query.order_by(Donor.full_name), page=page, per_page=DONORS_PER_PAGE, error_out=False
    )

    # Compute totals with one aggregate query instead of the N+1 pattern of
    # calling donor.total_donated / donor.donation_count (each of which
    # issues its own query) for every row on the page.
    donor_ids = [d.id for d in pagination.items]
    totals = {}
    if donor_ids:
        rows = (
            db.session.query(
                Donation.donor_id,
                func.coalesce(func.sum(Donation.amount), 0),
                func.count(Donation.id),
            )
            .filter(Donation.status == "success", Donation.donor_id.in_(donor_ids))
            .group_by(Donation.donor_id)
            .all()
        )
        totals = {donor_id: (float(total), count) for donor_id, total, count in rows}

    preachers_list = Preacher.query.order_by(Preacher.name).all()
    return render_template(
        "admin/donors.html", donors=pagination.items, pagination=pagination, totals=totals, q=q,
        donor_type=donor_type, preacher_id=preacher_id, donation_frequency=donation_frequency,
        preachers=preachers_list, donor_types=DONOR_TYPES, donor_type_labels=DONOR_TYPE_LABELS,
        donation_frequencies=DONATION_FREQUENCIES, donation_frequency_labels=DONATION_FREQUENCY_LABELS,
    )


@bp.route("/donors/<int:donor_id>")
@login_required
def donor_detail(donor_id):
    donor = Donor.query.get_or_404(donor_id)
    donations = donor.donations.order_by(Donation.donation_date.desc()).all()
    available_fys = sorted({d.financial_year for d in donations if d.financial_year and d.status == "success"}, reverse=True)
    return render_template(
        "admin/donor_detail.html", donor=donor, donations=donations, available_fys=available_fys,
        donor_type_labels=DONOR_TYPE_LABELS, donation_frequency_labels=DONATION_FREQUENCY_LABELS,
    )


def _parse_optional_date(form, key):
    """Parses an optional <input type=date> field (YYYY-MM-DD). Returns
    None for blank or malformed input rather than raising -- these are
    all "nice to have" relationship-building fields (DOB, anniversary,
    etc.), not required data, so a typo'd date shouldn't block saving the
    rest of the donor's details."""
    raw = (form.get(key) or "").strip()
    if not raw:
        return None
    try:
        return datetime.datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError:
        return None


@bp.route("/donors/<int:donor_id>/edit", methods=["GET", "POST"])
@login_required
def donor_edit(donor_id):
    donor = Donor.query.get_or_404(donor_id)
    if request.method == "POST":
        form = request.form
        pan = form.get("pan", "").strip().upper()
        if pan and not is_valid_pan(pan):
            flash("That PAN doesn't look right. It should be 10 characters like ABCDE1234F.")
            return redirect(url_for("admin.donor_edit", donor_id=donor.id))

        if not is_valid_phone(form.get("phone")):
            flash("That phone number doesn't look right. Please enter a 10-digit mobile number.")
            return redirect(url_for("admin.donor_edit", donor_id=donor.id))
        if not is_valid_phone(form.get("whatsapp_number")):
            flash("That WhatsApp number doesn't look right. Please enter a 10-digit mobile number.")
            return redirect(url_for("admin.donor_edit", donor_id=donor.id))

        donor.full_name = form.get("full_name", "").strip() or donor.full_name
        # normalize_phone() collapses "+91 88020 81265" / "918802081265" /
        # "08802081265" / "8802081265" down to the same plain 10-digit
        # value this app stores/matches everywhere -- otherwise editing a
        # donor's number into a different-but-equivalent format here would
        # silently break donor login and future donation dedup matching.
        donor.phone = normalize_phone(form.get("phone")) or None
        donor.whatsapp_number = normalize_phone(form.get("whatsapp_number")) or None
        donor.email = form.get("email", "").strip().lower() or None
        donor.pan = pan or None
        donor.address = form.get("address", "").strip() or None
        donor.city = form.get("city", "").strip() or None
        donor.state = form.get("state", "").strip() or None
        donor.pincode = form.get("pincode", "").strip() or None

        # -- Relationship-management fields --
        donor_type = form.get("donor_type") or ""
        donor.donor_type = donor_type if donor_type in DONOR_TYPES else None

        donation_frequency = form.get("donation_frequency") or ""
        donor.donation_frequency = donation_frequency if donation_frequency in DONATION_FREQUENCIES else None

        preacher_id = None
        raw_preacher_id = form.get("connected_preacher_id")
        if raw_preacher_id:
            try:
                candidate = int(raw_preacher_id)
                if Preacher.query.get(candidate):
                    preacher_id = candidate
            except ValueError:
                pass
        donor.connected_preacher_id = preacher_id

        donor.gifts = (form.get("gifts") or "").strip()[:500] or None
        donor.additional_info = (form.get("additional_info") or "").strip() or None
        donor.dob = _parse_optional_date(form, "dob")
        donor.father_dob = _parse_optional_date(form, "father_dob")
        donor.mother_dob = _parse_optional_date(form, "mother_dob")
        donor.wife_dob = _parse_optional_date(form, "wife_dob")
        donor.marriage_anniversary = _parse_optional_date(form, "marriage_anniversary")

        log_activity("donor_edit", target_type="donor", target_id=donor.id, details=f"Edited donor '{donor.full_name}'")
        db.session.commit()
        flash("Donor details updated.")
        return redirect(url_for("admin.donor_detail", donor_id=donor.id))

    preachers_list = Preacher.query.filter_by(is_active=True).order_by(Preacher.name).all()
    if donor.connected_preacher and not donor.connected_preacher.is_active:
        # Keep the currently-assigned preacher selectable even if they've
        # since been deactivated -- otherwise saving this form with no
        # other changes would silently clear the assignment.
        preachers_list = sorted(preachers_list + [donor.connected_preacher], key=lambda p: p.name)
    return render_template(
        "admin/donor_edit.html", donor=donor, preachers=preachers_list,
        donor_types=DONOR_TYPES, donor_type_labels=DONOR_TYPE_LABELS,
        donation_frequencies=DONATION_FREQUENCIES, donation_frequency_labels=DONATION_FREQUENCY_LABELS,
    )


@bp.route("/donors/<int:donor_id>/merge", methods=["POST"])
@login_required
@admin_role_required
def donor_merge(donor_id):
    """Merges a duplicate donor record into this one: reassigns all of the
    duplicate's donations, backfills any blank fields on the kept record,
    then removes the duplicate. This is the fix for donors who ended up
    with two records (e.g. a typo'd phone number) despite the dedup logic
    on the donation form."""
    keep = Donor.query.get_or_404(donor_id)
    lookup = request.form.get("duplicate_lookup", "").strip()

    if not lookup:
        flash("Enter the duplicate donor's phone or email to merge.")
        return redirect(url_for("admin.donor_detail", donor_id=donor_id))

    # Accept the phone half of the lookup in any format (with/without
    # +91, spaces, leading 0) -- same normalize_phone() used everywhere
    # else a phone number is entered or matched against.
    duplicate = Donor.query.filter(
        Donor.id != keep.id, (Donor.phone == normalize_phone(lookup)) | (Donor.email == lookup.lower())
    ).first()

    if duplicate is None:
        flash(f"No other donor found matching '{lookup}'.")
        return redirect(url_for("admin.donor_detail", donor_id=donor_id))

    moved = Donation.query.filter_by(donor_id=duplicate.id).update({"donor_id": keep.id})
    keep.full_name = keep.full_name or duplicate.full_name
    keep.phone = keep.phone or duplicate.phone
    keep.email = keep.email or duplicate.email
    keep.pan = keep.pan or duplicate.pan
    keep.address = keep.address or duplicate.address
    keep.city = keep.city or duplicate.city
    keep.state = keep.state or duplicate.state
    keep.pincode = keep.pincode or duplicate.pincode

    db.session.delete(duplicate)
    log_activity(
        "donor_merge", target_type="donor", target_id=keep.id,
        details=f"Merged duplicate donor #{duplicate.id} ('{duplicate.full_name}') into '{keep.full_name}' ({moved} donation(s) reassigned)",
    )
    db.session.commit()

    flash(f"Merged {moved} donation(s) from the duplicate record into {keep.full_name}.")
    return redirect(url_for("admin.donor_detail", donor_id=donor_id))


@bp.route("/donors/duplicates")
@login_required
def duplicate_donors():
    """Donors who share a phone number or email but have different names on
    file -- worth a quick human check.

    find_or_create_donor (public.py) only treats a phone/email match as
    "the same donor" when the name also agrees -- a shared family phone
    number no longer lets one donation silently overwrite a different
    family member's name/PAN/address. The flip side is that every group
    listed here is, by design, already kept as separate donor records; this
    page exists so staff can glance through them and confirm that's
    correct (genuinely different people sharing one contact) rather than
    two records for the same person (a nickname vs. legal name, a retyped
    name that didn't quite match, etc). Use the merge tool on a donor's own
    page (enter the other one's phone or email) to combine two records that
    really are the same person.
    """
    phone_groups = defaultdict(list)
    for donor in Donor.query.filter(Donor.phone.isnot(None)).order_by(Donor.full_name).all():
        phone_groups[donor.phone].append(donor)
    phone_dupes = sorted(
        (group for group in phone_groups.values() if len(group) > 1),
        key=lambda g: g[0].phone,
    )

    email_groups = defaultdict(list)
    for donor in Donor.query.filter(Donor.email.isnot(None)).order_by(Donor.full_name).all():
        email_groups[donor.email].append(donor)
    email_dupes = sorted(
        (group for group in email_groups.values() if len(group) > 1),
        key=lambda g: g[0].email,
    )

    return render_template(
        "admin/duplicate_donors.html",
        phone_dupes=phone_dupes, email_dupes=email_dupes,
    )


def _apply_donations_filters(query):
    """Shared campaign/mode/status/date-range filtering for the Donations
    Log page and its CSV export, kept in one place so the export can never
    drift out of sync with what's actually on screen (it's meant to be
    "everything currently filtered", just without pagination).

    Returns (query, filters) where filters is a dict of the resolved
    values, ready to splice straight into the template context.
    """
    # Defaults to "success" (the historical behaviour every other caller of
    # this route already relies on) but a status=... query param lets staff
    # pull up cancelled donations too, or "all" to see everything mixed
    # together -- otherwise a cancelled donation would just silently vanish
    # from this list with no way to find it again.
    status = request.args.get("status", "success")
    if status != "all":
        query = query.filter_by(status=status)

    campaign_id = request.args.get("campaign_id", type=int)
    if campaign_id:
        query = query.filter_by(campaign_id=campaign_id)

    mode = request.args.get("mode")
    if mode:
        query = query.filter_by(payment_mode=mode)

    date_from_raw = request.args.get("date_from") or ""
    date_to_raw = request.args.get("date_to") or ""

    # Quick ranges. The log is overwhelmingly used to answer "what came in
    # recently", and typing two dates for that every time is friction on
    # the most common task -- so a fresh visit lands on the current month
    # rather than the full history.
    #
    # Explicit dates always win: a link carrying date_from/date_to is a
    # deliberate custom range (including one this very form produced), and
    # must not be quietly overridden by a preset.
    date_range = request.args.get("range") or ""
    if date_from_raw or date_to_raw:
        date_range = "custom"
    elif not date_range:
        # No dates and no preset -- the fresh-visit case. Anything wanting
        # the old unbounded behaviour asks for range=all explicitly (see
        # the dashboard's pending/failed links, where an older donation is
        # precisely what staff are looking for).
        date_range = "this_month"

    if date_range not in ("custom", "all"):
        today = datetime.date.today()
        start = end = None
        if date_range == "this_month":
            start, end = today.replace(day=1), today
        elif date_range == "last_month":
            end = today.replace(day=1) - datetime.timedelta(days=1)
            start = end.replace(day=1)
        elif date_range == "last_3_months":
            # Three whole calendar months including this one, not a rolling
            # 90 days -- staff reconcile by month, not by day.
            month, year = today.month - 2, today.year
            if month <= 0:
                month, year = month + 12, year - 1
            start, end = datetime.date(year, month, 1), today
        elif date_range == "this_fy":
            # Indian financial year, 1 April - 31 March: the same boundary
            # receipt numbering and the Form 10BD filing already use.
            fy_start_year = today.year if today.month >= 4 else today.year - 1
            start, end = datetime.date(fy_start_year, 4, 1), today
        else:
            date_range = "all"  # unrecognised value -- don't filter on a guess

        if start:
            date_from_raw, date_to_raw = start.isoformat(), end.isoformat()

    try:
        if date_from_raw:
            date_from = datetime.datetime.strptime(date_from_raw, "%Y-%m-%d")
            query = query.filter(Donation.donation_date >= date_from)
        if date_to_raw:
            date_to = datetime.datetime.strptime(date_to_raw, "%Y-%m-%d")
            # donation_date carries a real time-of-day for online donations
            # (see below), so a plain <= comparison against midnight of the
            # "to" date would silently exclude every online donation made
            # later that same day. Push the upper bound to the start of the
            # next day instead.
            query = query.filter(Donation.donation_date < date_to + datetime.timedelta(days=1))
    except ValueError:
        date_from_raw = date_to_raw = ""

    return query, {
        "status": status, "campaign_id": campaign_id, "mode": mode,
        "date_from": date_from_raw, "date_to": date_to_raw,
        "date_range": date_range,
    }


# Column-header sorting for the Donations Log. Sort key is a plain column
# name ("amount") for ascending, or the same name prefixed with "-" for
# descending ("-amount") -- mirrors the convention already used by
# Django/DRF-style APIs so it reads naturally in a URL. Donor/Campaign
# sort by the related table's name rather than the foreign key id, which
# needs a join.
_DONATIONS_SORT_COLUMNS = {"date", "donor", "campaign", "amount", "status"}


def _sorted_donations_query(query, sort_key):
    descending = sort_key.startswith("-")
    key = sort_key[1:] if descending else sort_key

    if key not in _DONATIONS_SORT_COLUMNS:
        # Unrecognised/typo'd sort param -- fall back to the historical
        # default rather than erroring the page out.
        key, descending = "date", True

    if key == "donor":
        query = query.join(Donor, Donation.donor_id == Donor.id)
        column = Donor.full_name
    elif key == "campaign":
        query = query.join(Campaign, Donation.campaign_id == Campaign.id)
        column = Campaign.name
    elif key == "amount":
        column = Donation.amount
    elif key == "status":
        column = Donation.status
    else:
        column = Donation.donation_date

    query = query.order_by(column.desc() if descending else column.asc())
    resolved_sort = ("-" if descending else "") + key
    return query, resolved_sort


@bp.route("/donations")
@login_required
def donations():
    query, filters = _apply_donations_filters(Donation.query)
    query, sort = _sorted_donations_query(query, request.args.get("sort", "-date"))
    page = request.args.get("page", 1, type=int)
    pagination = db.paginate(query, page=page, per_page=DONATIONS_PER_PAGE, error_out=False)
    campaigns = Campaign.query.order_by(Campaign.name).all()
    return render_template(
        "admin/donations.html", donations=pagination.items, pagination=pagination, campaigns=campaigns,
        sort=sort, **filters,
    )


@bp.route("/donations/<int:donation_id>/cancel", methods=["POST"])
@login_required
@admin_role_required
def cancel_donation(donation_id):
    """Cancels a donation/receipt rather than deleting it. Sets status to
    "cancelled" instead of a separate boolean flag so it's automatically
    excluded from every existing money-total query, export, and report
    that already filters on status == "success" (dashboard totals, donor
    totals, campaign totals, Form 10BD/collections exports, lapsed-donor
    list) -- with zero changes needed at any of those call sites.

    The receipt PDF itself is never touched -- receipts are immutable once
    issued in this system. The /receipt/<id> download route already
    refuses to serve non-"success" donations, so a cancelled receipt just
    becomes undownloadable rather than being altered or reissued.
    """
    donation = Donation.query.get_or_404(donation_id)
    if donation.status == "cancelled":
        flash("That donation is already cancelled.")
        return redirect(url_for("admin.donor_detail", donor_id=donation.donor_id))
    if donation.status != "success":
        flash("Only successful donations can be cancelled.")
        return redirect(url_for("admin.donor_detail", donor_id=donation.donor_id))

    reason = (request.form.get("cancellation_reason") or "").strip()[:300]
    if not reason:
        flash("Please give a reason for cancelling this donation.")
        return redirect(url_for("admin.donor_detail", donor_id=donation.donor_id))

    donation.status = "cancelled"
    donation.cancelled_at = datetime.datetime.utcnow()
    donation.cancelled_by = current_user.username
    donation.cancellation_reason = reason
    log_activity(
        "donation_cancel", target_type="donation", target_id=donation.id,
        details=f"Cancelled donation {donation.receipt_number or ('#' + str(donation.id))} (Rs. {donation.amount}) -- reason: {reason}",
    )
    db.session.commit()

    flash(f"Donation {donation.receipt_number or ('#' + str(donation.id))} has been cancelled.")
    return redirect(url_for("admin.donor_detail", donor_id=donation.donor_id))


@bp.route("/donations/<int:donation_id>/restore", methods=["POST"])
@login_required
@admin_role_required
def restore_donation(donation_id):
    """Undoes an accidental cancellation -- restores status to "success"
    and clears the cancellation fields. The donation reappears in every
    total/export it had been excluded from, and its receipt becomes
    downloadable again."""
    donation = Donation.query.get_or_404(donation_id)
    if donation.status != "cancelled":
        flash("That donation isn't cancelled.")
        return redirect(url_for("admin.donor_detail", donor_id=donation.donor_id))

    donation.status = "success"
    donation.cancelled_at = None
    donation.cancelled_by = None
    donation.cancellation_reason = None
    log_activity(
        "donation_restore", target_type="donation", target_id=donation.id,
        details=f"Restored donation {donation.receipt_number or ('#' + str(donation.id))} (Rs. {donation.amount})",
    )
    db.session.commit()

    flash(f"Donation {donation.receipt_number or ('#' + str(donation.id))} has been restored.")
    return redirect(url_for("admin.donor_detail", donor_id=donation.donor_id))


@bp.route("/donations/<int:donation_id>/finalize-pending", methods=["POST"])
@login_required
@admin_role_required
def finalize_pending_donation(donation_id):
    """Manually finalizes an online donation stuck in "pending" even though
    the donor's money was actually deducted -- covers the rare case where
    the webhook never fired (misconfigured, down, or the request never made
    it through) *and* the donor's browser never got a chance to report back
    either (tab/app closed right after paying, connection dropped). Without
    this there was no way to issue that donor a receipt short of a direct
    database edit.

    Requires the actual Razorpay Payment ID (starts "pay_...", found on the
    Razorpay Dashboard -> Payments, searchable by amount/phone/time) typed
    in by staff -- this isn't a "trust the donor" override, it's staff
    confirming against Razorpay's own records that the payment genuinely
    went through before a receipt gets issued for it.
    """
    donation = Donation.query.get_or_404(donation_id)
    if donation.payment_mode != "online" or donation.status != "pending":
        flash("This action only applies to a still-pending online donation.")
        return redirect(url_for("admin.donations"))

    payment_id = (request.form.get("razorpay_payment_id") or "").strip()
    if not payment_id:
        flash("Enter the Razorpay Payment ID (from the Razorpay Dashboard -> Payments) to confirm this payment actually went through.")
        return redirect(url_for("admin.donations"))

    donation.razorpay_payment_id = payment_id
    _finalize_success(donation)
    log_activity(
        "donation_manual_finalize", target_type="donation", target_id=donation.id,
        details=f"Manually finalized pending donation as paid (Payment ID {payment_id}), receipt {donation.receipt_number}",
    )
    db.session.commit()

    flash(f"Donation marked as paid -- receipt {donation.receipt_number} generated and sent.")
    return redirect(url_for("admin.donations"))


def _validated_id_from_form(form, key, model, label):
    """Same idea as public.py's _validated_fk_id, but for the admin
    manual-donation form (werkzeug form data, and flash+redirect instead of
    a JSON error response on failure)."""
    raw = form.get(key)
    if not raw:
        return None, None
    try:
        value = int(raw)
    except (TypeError, ValueError):
        return None, f"Invalid {label}."
    if not model.query.get(value):
        return None, f"Invalid {label}."
    return value, None


def _offline_donation_form_context():
    """Dropdown data the Single Entry tab needs -- shared between
    manual_donation() and bulk_import_donations() since both routes can
    render the merged offline_donation.html page (Single Entry + Bulk
    Upload live on one page/one nav tab now)."""
    return {
        "campaigns": Campaign.query.filter_by(is_active=True).order_by(Campaign.name).all(),
        "bace_properties": BaceProperty.query.filter_by(is_active=True).order_by(BaceProperty.name).all(),
        "festivals": Festival.query.filter_by(is_active=True).order_by(Festival.name).all(),
        "seva_types": SevaType.query.filter_by(is_active=True).order_by(SevaType.name).all(),
        "live_to_give_purposes": LiveToGivePurpose.query.filter_by(is_active=True).order_by(LiveToGivePurpose.name).all(),
        # now_ist(), not datetime.date.today() -- see the same fix applied
        # throughout this file for "today"-as-server-clock bugs. This one
        # only sets the Donation Date field's default value (staff can
        # still change it), but during the ~5.5 IST hours a day where the
        # server's UTC clock hasn't rolled over yet, the un-fixed version
        # would default the field to yesterday's date.
        "today": now_ist().date(),
    }


# Column widths taken directly from the Donor model (models.py) -- kept
# here as a single source of truth for the offline-donation entry points,
# so a value that's too long for its column never reaches the database as
# an unhandled sqlalchemy.exc.DataError (which, mid-transaction, can leave
# the DB session broken for whatever happens next in the same request).
# The single-entry form's Address/City/State/Remarks fields have no
# client-side maxlength, and CSV import rows are entirely unvalidated
# free text, so this is the actual line of defence, not a formality.
_DONOR_FIELD_LIMITS = {"full_name": 200, "email": 200, "pan": 10, "address": 400, "city": 100, "state": 100, "pincode": 10}


def _sanitize_donor_data(data):
    """Builds the plain dict find_or_create_donor()/Donor(...) expect,
    trimmed and clipped to each field's real column width. `data` can be
    a werkzeug form (single-entry POST) or a CSV row dict (bulk import) --
    either way, only .get() is used, so both work identically here."""
    out = {
        key: (data.get(key) or "").strip()
        for key in ("full_name", "email", "pan", "address", "city", "state", "pincode")
    }
    for key, limit in _DONOR_FIELD_LIMITS.items():
        out[key] = out[key][:limit]
    # Not length-clipped: normalize_phone() (called inside
    # find_or_create_donor) reduces these to a bounded 10-digit string,
    # and is_valid_phone() has already rejected anything malformed before
    # this function is ever called.
    out["phone"] = data.get("phone")
    out["whatsapp_number"] = data.get("whatsapp_number")
    return out


class _UploadReadError(Exception):
    """The upload couldn't be read at all. Carries a message meant for the
    person who uploaded it, not a stack trace."""


class _UploadedTable:
    """Rows from an uploaded .csv or .xlsx, in one shape.

    Presents the same surface the importers already used from
    csv.DictReader -- `.fieldnames` and iteration yielding dicts -- so
    reading a spreadsheet is a change of source, not a change to four
    import routines.
    """

    def __init__(self, fieldnames, rows):
        self.fieldnames = fieldnames
        self._rows = rows

    def __iter__(self):
        return iter(self._rows)


def _excel_cell_to_text(value):
    """One spreadsheet cell as the string an importer expects.

    This is where the value of accepting .xlsx actually lands. In a CSV
    everything has already been flattened to text by whatever wrote it,
    and the damage is done: a date has become 01/08/2026 with no record of
    which number was the month, 1100 has become "1,100", and a phone
    number has become 9.87654e+09. In an .xlsx the cell still knows what
    it is, so a date arrives as a date and is written back out in the
    canonical form with nothing guessed.
    """
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, float):
        # Excel stores every number as a float, so an amount of 1100 and a
        # phone number both arrive as 1100.0 / 9876543210.0. Render whole
        # numbers without the ".0" -- a phone number ending in ".0" fails
        # validation, and str() on a large float would reach for
        # scientific notation, which is exactly the mangling this is
        # meant to avoid.
        if value == int(value):
            return str(int(value))
        return repr(value)
    return str(value).strip()


def _table_from_upload(file):
    """Rows from an uploaded .csv or .xlsx file.

    CSV is read into memory and decoded rather than wrapping the raw
    stream in io.TextIOWrapper. Werkzeug hands uploads over as a
    SpooledTemporaryFile, which does not implement readable() before
    Python 3.11 -- TextIOWrapper requires it, so the wrap raises
    AttributeError and the route reports the file as unreadable.
    Production runs 3.12 and never saw it; every CSV import was silently
    broken on older local Pythons, which is where they get tried first.

    Decoding with errors="replace" rather than strict: a single stray byte
    from a spreadsheet export shouldn't make a whole file unreadable, and
    the per-row validation will flag anything that actually matters.

    These are staff-uploaded imports of at most a few thousand rows, so
    reading them into memory is not a concern either way.
    """
    raw = file.read()
    if not isinstance(raw, bytes):
        raw = (raw or "").encode("utf-8")

    # .xlsx is a zip archive, so it always starts "PK". Sniffing the
    # content rather than trusting the extension: a file saved as
    # "donations.csv" from Excel's xlsx format is a real mistake people
    # make, and it produces binary gibberish through the CSV path.
    looks_like_xlsx = raw[:2] == b"PK"
    named_xlsx = (file.filename or "").lower().endswith((".xlsx", ".xlsm"))

    # Old-format .xls is an OLE2 compound document, not a zip, and openpyxl
    # can't read it. Worth naming: through the CSV path it decodes to
    # binary noise and the operator is told their file is missing every
    # required column, which sends them looking in the wrong place.
    if raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        raise _UploadReadError(
            "That's an older Excel .xls file, which can't be read directly. Open it in "
            "Excel and use Save As to make it an .xlsx (or a CSV), then upload that."
        )

    if looks_like_xlsx or named_xlsx:
        if not looks_like_xlsx:
            raise _UploadReadError(
                "That file is named like an Excel workbook but isn't one. Re-save it "
                "from Excel as .xlsx, or export it as CSV."
            )
        return _table_from_xlsx(raw)

    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return _UploadedTable(list(reader.fieldnames or []), list(reader))


def _table_from_xlsx(raw):
    """Rows from an .xlsx workbook: first sheet, first row as headers."""
    try:
        import openpyxl
    except ImportError:
        raise _UploadReadError(
            "Excel (.xlsx) uploads need the openpyxl package, which isn't installed "
            "here. Run `pip install -r requirements.txt`, or save the file as CSV "
            "and upload that instead."
        )

    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(raw), data_only=True, read_only=True)
    except Exception:
        raise _UploadReadError(
            "That Excel file couldn't be opened. If it's an older .xls, open it in "
            "Excel and save as .xlsx, or export it as CSV."
        )

    try:
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)

        headers = None
        for row in rows:
            values = [_excel_cell_to_text(v) for v in row]
            if any(v for v in values):  # skip blank leading rows
                headers = [v.strip() for v in values]
                break
        if not headers:
            raise _UploadReadError("That Excel file has no column headings in it.")

        # Trailing blank columns are what Excel leaves behind after
        # someone deletes a column, and an unnamed header would collide in
        # the row dict.
        keep = [i for i, h in enumerate(headers) if h]
        headers = [headers[i] for i in keep]

        parsed = []
        for row in rows:
            values = [_excel_cell_to_text(v) for v in row]
            if not any(v for v in values):
                continue  # blank row -- Excel files are full of them
            parsed.append({
                headers[n]: (values[i] if i < len(values) else "")
                for n, i in enumerate(keep)
            })
        return _UploadedTable(headers, parsed)
    finally:
        workbook.close()


def _normalize_camp_text(value):
    """Trim and collapse internal whitespace on a camp or batch name.

    Camp/batch are plain text and reports group by the exact string, so
    "Utkarsha  2026 " and "Utkarsha 2026" would otherwise be two separate
    camps in every total. This can only fix whitespace -- a real
    misspelling still splits the total until the rows are edited -- but
    whitespace is far and away the most common way a pasted or exported
    value differs from a typed one, and it costs nothing to rule out.

    Case is deliberately left alone: camp names carry meaningful
    capitalisation ("BACE", "IYF") that title-casing would mangle.
    """
    if not value:
        return None
    return " ".join(str(value).split())[:150] or None


IYF_CAMP_CAMPAIGN_NAME = "IYF Camps"


def _iyf_camp_campaign(create=True):
    """The campaign every camp collection is recorded against.

    Donation.campaign_id is required, and camp money is its own stream --
    filing it under an existing campaign would corrupt that campaign's
    totals. Created on first use rather than needing a setup step, and
    is_80g=False per the rule that camp collections aren't 80G-eligible.

    `create=False` returns None instead of creating it, and exists for the
    bulk import's preview: this function commits, so a dry run that called
    it was writing a Campaign row to the database while telling the
    operator nothing had been saved. Harmless in itself -- the row is
    identical to the one a real import would make -- but "nothing has been
    saved" has to be true, or it isn't worth saying.
    """
    campaign = Campaign.query.filter_by(name=IYF_CAMP_CAMPAIGN_NAME).first()
    if campaign is None and create:
        campaign = Campaign(
            name=IYF_CAMP_CAMPAIGN_NAME,
            description="Donations collected from students at IYF camps.",
            is_80g=False,
        )
        db.session.add(campaign)
        db.session.commit()
    return campaign


def _known_batch_names():
    """Batch names already in use, for the entry form's picker.

    Batches stay free text -- they're per-camp, short-lived, and there are
    far too many to be worth maintaining as records. Offering the ones
    already used is enough to keep spelling consistent in practice.
    """
    return sorted(
        b for (b,) in db.session.query(Donation.batch_name)
        .filter(Donation.batch_name.isnot(None)).distinct().all() if b
    )


def _resolve_camp_name(raw):
    """Match a camp name from an import against the Camp list.

    Case- and whitespace-insensitive, returning the camp's own spelling so
    that's what lands on the donation. A Zoho export writing "utkarsha
    2026" therefore files under "Utkarsha 2026" rather than starting a
    second camp that differs only in case.

    Returns None if there's no such camp. Deliberately does not create one:
    the point of a managed list is that an unrecognised name is a typo to
    be caught, not a new camp to be silently invented.

    Matches inactive camps too -- retiring a camp shouldn't break an import
    of donations that were collected while it was running.
    """
    cleaned = _normalize_camp_text(raw)
    if not cleaned:
        return None
    camp = Camp.query.filter(func.lower(Camp.name) == cleaned.lower()).first()
    return camp.name if camp else None


def _payment_reference_error(payment_mode, cheque_number=None, bank_transaction_id=None):
    """Every payment except cash has to arrive with its reference.

    Cash is the only mode with nothing to reference -- it happened at the
    counter and the receipt is the record. Everything else already exists
    in someone else's system (a bank statement, a cheque, Razorpay, Zoho),
    and the reference is the only way to match this donation to it when
    the two are reconciled. Recorded without one, the row is unverifiable
    forever, and nobody goes back to fill it in.

    Returns an error string, or None if the entry is acceptable.

    Either field satisfies it. The camp form collects a cheque's reference
    as bank_transaction_id (it has no separate cheque-number field), so
    requiring the mode's "own" field would reject entries that do carry a
    reference.
    """
    mode = (payment_mode or "").strip().lower()
    if mode == "cash":
        return None
    if (cheque_number or "").strip() or (bank_transaction_id or "").strip():
        return None
    if mode == "cheque":
        return (
            "A cheque donation needs its cheque number -- it's what ties this "
            "receipt to the cheque when it clears."
        )
    label = {
        "bank_transfer": "A bank transfer needs its UTR / transaction ID",
        "online": "An online payment needs its transaction / payment ID",
    }.get(mode, f"A {mode.replace('_', ' ') or 'non-cash'} donation needs its reference")
    return (
        f"{label} -- it's the only way to match this donation to the payment "
        "later. Only cash can be recorded without a reference."
    )


def _create_offline_donation(
    *, donor_data, campaign, amount, payment_mode, donation_date, recorded_by,
    bace_property_id=None, festival_id=None, seva_type_id=None, live_to_give_purpose_id=None,
    is_80g_requested=None, cheque_number=None, cheque_bank_name=None, bank_transaction_id=None,
    remarks=None, send_notifications=True, camp_name=None, batch_name=None,
):
    """The one shared path every offline donation -- single-entry form or
    a bulk CSV row -- goes through once its own field-by-field validation
    has already passed. Used to be two near-identical copies of this
    logic (one in manual_donation(), one in bulk_import_donations()),
    which is exactly how the two drifted out of sync before: bulk import
    truncated `remarks` to fit its DB column, the single-entry form
    didn't. One implementation now, used by both.

    Three stages, each safe to fail on its own:
      1. Find/create the donor + create the Donation row + issue a real
         receipt number, in one DB transaction. If *anything* in this
         stage raises -- a bad value the earlier validation missed, a
         transient DB error, whatever -- it's caught, the transaction is
         rolled back, and nothing is left half-written. This stage used
         to be completely unguarded in manual_donation(), which meant any
         exception here propagated straight out of the view function.
      2. Generate the receipt PDF and store it on the (already-committed)
         donation. A failure here no longer loses the donation record or
         its receipt number -- it's reported back as `pdf_ok: False` so
         the caller can say so, rather than crashing.
      3. Kick off the background email/WhatsApp notification thread
         (never inline -- see the comment at the call site below for why).

    Returns a dict:
      {"ok": False, "error": "..."} -- nothing was saved, stage 1 failed.
      {"ok": True, "donor": Donor, "donation": Donation,
       "receipt_number": str, "pdf_ok": bool} -- stage 1 always
      succeeded if "ok" is True; "pdf_ok" says whether stages 2/3 did too.
    """
    reference_error = _payment_reference_error(payment_mode, cheque_number, bank_transaction_id)
    if reference_error:
        # Checked before find_or_create_donor, so a rejected row doesn't
        # leave a new donor behind for a donation that was never created.
        #
        # Enforced here rather than in each caller because here is the one
        # place all four offline entry points meet: the single-entry form,
        # the bulk CSV import, IYF camp single entry and IYF camp bulk. A
        # rule written into the forms would have to be remembered four
        # times, and again by the fifth form somebody adds.
        #
        # import_legacy_donations builds its Donation rows directly and so
        # isn't subject to this -- deliberately. It backfills records from
        # before this system, whose references were never captured: of the
        # ~6,000 rows staged for go-live, 3,509 have none. Requiring one
        # there would reject the temple's own history in order to enforce a
        # rule about how payments are recorded from now on.
        return {"ok": False, "error": reference_error}

    try:
        donor = find_or_create_donor(_sanitize_donor_data(donor_data))

        donation = Donation(
            donor_id=donor.id,
            campaign_id=campaign.id,
            amount=amount,
            payment_mode=payment_mode,
            status="success",
            donation_date=donation_date,
            bace_property_id=bace_property_id,
            festival_id=festival_id,
            seva_type_id=seva_type_id,
            live_to_give_purpose_id=live_to_give_purpose_id,
            is_80g_requested=is_80g_requested,
            cheque_number=(cheque_number or "").strip()[:50] or None,
            cheque_bank_name=(cheque_bank_name or "").strip()[:150] or None,
            bank_transaction_id=(bank_transaction_id or "").strip()[:100] or None,
            camp_name=_normalize_camp_text(camp_name),
            batch_name=_normalize_camp_text(batch_name),
            remarks=(remarks or "").strip()[:300] or None,
            recorded_by=recorded_by,
        )
        db.session.add(donation)
        db.session.flush()

        receipt_number, fy = ReceiptCounter.next_receipt_number(donation.effective_is_80g, donation_date)
        donation.receipt_number = receipt_number
        donation.financial_year = fy
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception(
            "Failed to create offline donation (campaign_id=%s, amount=%s, recorded_by=%s)",
            campaign.id, amount, recorded_by,
        )
        return {"ok": False, "error": f"Couldn't save the donation ({exc}). Nothing was recorded -- please try again."}

    result = {"ok": True, "donor": donor, "donation": donation, "receipt_number": receipt_number, "pdf_ok": True}

    try:
        pdf_bytes = generate_receipt_pdf(donation, donor, campaign, _org_cfg())
        donation.receipt_pdf = pdf_bytes
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception(
            "Receipt PDF generation failed for donation id=%s (receipt %s) -- donation record was saved, "
            "PDF/notifications were not.", donation.id, receipt_number,
        )
        result["pdf_ok"] = False
        return result

    if send_notifications:
        # Backgrounded, not sent inline -- same reasoning as public.py's
        # _finalize_success(): a slow/hanging SMTP connection or WhatsApp
        # API call stacked on top of PDF generation can blow past
        # gunicorn's worker timeout, which kills the whole worker mid-
        # response (a dropped connection in the browser, no clean error,
        # nothing in the logs) rather than a normal exception. The
        # receipt is already saved either way by this point.
        try:
            app = current_app._get_current_object()
            if app.config.get("TESTING"):
                _send_receipt_notifications_background(app, donation.id, pdf_bytes)
            else:
                threading.Thread(
                    target=_send_receipt_notifications_background, args=(app, donation.id, pdf_bytes), daemon=True
                ).start()
        except Exception:
            current_app.logger.exception(
                "Failed to start receipt notification thread for donation id=%s (receipt %s).",
                donation.id, receipt_number,
            )

    return result


@bp.route("/donations/manual", methods=["GET", "POST"])
@login_required
def manual_donation():
    if request.method == "POST":
        form = request.form

        try:
            campaign_id = int(form["campaign_id"])
            amount = float(form["amount"])
        except (KeyError, TypeError, ValueError):
            flash("Please choose a campaign and enter a valid amount.")
            return redirect(url_for("admin.manual_donation"))
        campaign = Campaign.query.get_or_404(campaign_id)

        pan = (form.get("pan") or "").strip()
        if pan and not is_valid_pan(pan):
            flash("That PAN doesn't look right. It should be 10 characters like ABCDE1234F.")
            return redirect(url_for("admin.manual_donation"))

        if not is_valid_phone(form.get("phone")):
            flash("That phone number doesn't look right. Please enter a 10-digit mobile number.")
            return redirect(url_for("admin.manual_donation"))
        if not is_valid_phone(form.get("whatsapp_number")):
            flash("That WhatsApp number doesn't look right. Please enter a 10-digit mobile number.")
            return redirect(url_for("admin.manual_donation"))

        high_value_error = high_value_pan_address_error(amount, pan, form.get("address"))
        if high_value_error:
            flash(high_value_error)
            return redirect(url_for("admin.manual_donation"))

        bace_property_id, error = _validated_id_from_form(form, "bace_property_id", BaceProperty, "BACE property")
        if error:
            flash(error)
            return redirect(url_for("admin.manual_donation"))
        festival_id, error = _validated_id_from_form(form, "festival_id", Festival, "festival")
        if error:
            flash(error)
            return redirect(url_for("admin.manual_donation"))
        seva_type_id, error = _validated_id_from_form(form, "seva_type_id", SevaType, "seva type")
        if error:
            flash(error)
            return redirect(url_for("admin.manual_donation"))
        live_to_give_purpose_id, error = _validated_id_from_form(
            form, "live_to_give_purpose_id", LiveToGivePurpose, "donation purpose"
        )
        if error:
            flash(error)
            return redirect(url_for("admin.manual_donation"))

        receipt_type = form.get("receipt_type")
        if receipt_type == "80g":
            is_80g_requested = True
        elif receipt_type == "non80g":
            is_80g_requested = False
        else:
            is_80g_requested = None

        # Same hard rule as the public form -- a purpose's own 80G
        # eligibility (LiveToGivePurpose.is_80g) can't be overridden by
        # picking "80g" here. Donation.effective_is_80g would silently
        # correct it anyway, but catching it here means staff find out
        # immediately instead of a receipt quietly coming out Non-80G.
        if live_to_give_purpose_id and is_80g_requested:
            purpose = LiveToGivePurpose.query.get(live_to_give_purpose_id)
            if purpose and not purpose.is_80g:
                flash(f"'{purpose.name}' isn't 80G-eligible -- select \"No\" for the 80G receipt question, or a different purpose.")
                return redirect(url_for("admin.manual_donation"))

        # Offline payment reference details. Still captured whichever mode
        # is selected -- a cheque number typed in and then the mode changed
        # back to Cash shouldn't block submission, it just goes unused.
        # What is enforced (in _create_offline_donation, so every entry
        # point gets it) is the other direction: any mode except cash must
        # arrive with a reference in one of these fields.
        cheque_number = form.get("cheque_number")
        cheque_bank_name = form.get("cheque_bank_name")
        bank_transaction_id = form.get("bank_transaction_id")

        donation_date_str = form.get("donation_date")
        try:
            donation_date = (
                datetime.datetime.strptime(donation_date_str, "%Y-%m-%d")
                if donation_date_str
                else datetime.datetime.utcnow()
            )
        except ValueError:
            flash("That donation date doesn't look right.")
            return redirect(url_for("admin.manual_donation"))

        result = _create_offline_donation(
            donor_data=form,
            campaign=campaign,
            amount=amount,
            payment_mode=form.get("payment_mode", "cash"),
            donation_date=donation_date,
            recorded_by=current_user.username,
            bace_property_id=bace_property_id,
            festival_id=festival_id,
            seva_type_id=seva_type_id,
            live_to_give_purpose_id=live_to_give_purpose_id,
            is_80g_requested=is_80g_requested,
            cheque_number=cheque_number,
            cheque_bank_name=cheque_bank_name,
            bank_transaction_id=bank_transaction_id,
            remarks=form.get("remarks"),
        )
        if not result["ok"]:
            flash(result["error"], "danger")
            return redirect(url_for("admin.manual_donation"))

        if result["pdf_ok"]:
            flash(f"Donation recorded. Receipt {result['receipt_number']} generated.")
        else:
            flash(
                f"Donation recorded with receipt {result['receipt_number']}, but generating the PDF failed. "
                f"Check the donor's record and regenerate the receipt if needed.", "warning",
            )
        return redirect(url_for("admin.donor_detail", donor_id=result["donor"].id))

    active_tab = "bulk" if request.args.get("tab") == "bulk" else "single"
    return render_template(
        "admin/offline_donation.html", active_tab=active_tab,
        bulk_results=None, bulk_created=None, bulk_skipped=None,
        **_offline_donation_form_context(),
    )


# Columns the demo file offers and the importer reads. Only the first five
# are mandatory per row -- everything else is optional and left blank/None
# if not supplied.
BULK_IMPORT_REQUIRED_COLUMNS = ["full_name", "campaign_name", "amount", "payment_mode", "donation_date"]
BULK_IMPORT_COLUMNS = [
    "full_name", "phone", "whatsapp_number", "email", "pan", "address", "city", "state", "pincode",
    "campaign_name", "amount", "payment_mode", "donation_date",
    "cheque_number", "cheque_bank_name", "bank_transaction_id",
    "receipt_type", "bace_property_name", "festival_name", "seva_type_name", "live_to_give_purpose_name",
    "remarks",
]
BULK_IMPORT_DEMO_ROWS = [
    {
        "full_name": "Ramesh Kumar", "phone": "9876543210", "whatsapp_number": "", "email": "ramesh@example.com",
        "pan": "ABCDE1234F", "address": "12 MG Road", "city": "Delhi", "state": "Delhi", "pincode": "110001",
        "campaign_name": "General Donations", "amount": "1100", "payment_mode": "cash",
        "donation_date": "2026-04-15", "cheque_number": "", "cheque_bank_name": "", "bank_transaction_id": "",
        "receipt_type": "", "bace_property_name": "", "festival_name": "", "seva_type_name": "",
        "live_to_give_purpose_name": "", "remarks": "Monthly seva",
    },
    {
        "full_name": "Sita Devi", "phone": "9123456780", "whatsapp_number": "", "email": "",
        "pan": "", "address": "", "city": "", "state": "", "pincode": "",
        "campaign_name": "Temple Construction", "amount": "5000", "payment_mode": "cheque",
        "donation_date": "2026-04-20", "cheque_number": "123456", "cheque_bank_name": "HDFC Bank",
        "bank_transaction_id": "", "receipt_type": "", "bace_property_name": "", "festival_name": "",
        "seva_type_name": "", "live_to_give_purpose_name": "", "remarks": "",
    },
    {
        "full_name": "Amit Sharma", "phone": "", "whatsapp_number": "9988776655", "email": "amit@example.com",
        "pan": "FGHIJ5678K", "address": "", "city": "", "state": "", "pincode": "",
        "campaign_name": "Live To Give", "amount": "2100", "payment_mode": "bank_transfer",
        "donation_date": "2026-04-22", "cheque_number": "", "cheque_bank_name": "",
        "bank_transaction_id": "UTR2026042212345", "receipt_type": "80g", "bace_property_name": "",
        "festival_name": "", "seva_type_name": "",
        "live_to_give_purpose_name": "Temple Construction (मंदिर निर्माण के लिए)",
        "remarks": "",
    },
]


@bp.route("/donations/bulk-import/demo.csv")
@login_required
@admin_role_required
def bulk_import_demo_csv():
    """A ready-to-edit template CSV -- every column the importer understands,
    header row plus a few example rows covering cash/cheque/bank transfer
    and an 80G Live To Give donation."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=BULK_IMPORT_COLUMNS)
    writer.writeheader()
    writer.writerows(BULK_IMPORT_DEMO_ROWS)
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=offline_donations_demo.csv"},
    )


@bp.route("/donations/bulk-import/demo.xlsx")
@login_required
@admin_role_required
def bulk_import_demo_xlsx():
    """The same template as an Excel workbook -- see _demo_workbook."""
    return _demo_workbook(
        BULK_IMPORT_COLUMNS, BULK_IMPORT_DEMO_ROWS, "offline_donations_demo.xlsx")


# Columns whose cell *type* matters when the template is opened in Excel.
# A CSV template can't carry this: Excel decides for itself what each
# column is on open, and its guesses are the source of most import
# trouble. Handing out a real workbook is the only way to fix the types
# up front, which is the whole reason .xlsx uploads are worth having.
_DEMO_DATE_COLUMNS = {
    "donation_date", "dob", "father_dob", "mother_dob", "wife_dob",
    "marriage_anniversary",
}
_DEMO_NUMBER_COLUMNS = {"amount"}
# Forced to Text. Excel turns a phone number into a float and renders it
# as 9.87654e+09; it strips the leading zero off a pincode; and a PAN or a
# reference that happens to look numeric gets the same treatment.
_DEMO_TEXT_COLUMNS = {
    "phone", "whatsapp_number", "pan", "pincode", "cheque_number",
    "bank_transaction_id", "receipt_number",
}


def _demo_workbook(columns, rows, filename):
    """The same template as the demo CSV, as a real .xlsx.

    Dates are written as dates and the text columns are formatted as text,
    so a file filled in from this one arrives with nothing to guess at --
    which is the point of uploading a workbook rather than an export of
    one.
    """
    try:
        import openpyxl
    except ImportError:
        abort(503, "Excel templates need openpyxl -- run pip install -r requirements.txt")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Import"
    sheet.append(list(columns))
    for cell in sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for row in rows:
        values = []
        for column in columns:
            value = row.get(column, "")
            if column in _DEMO_DATE_COLUMNS and value:
                try:
                    value = datetime.datetime.strptime(value, "%Y-%m-%d").date()
                except (TypeError, ValueError):
                    pass  # leave whatever the demo row had; it's an example
            elif column in _DEMO_NUMBER_COLUMNS and value:
                try:
                    value = float(value)
                except (TypeError, ValueError):
                    pass
            values.append(value)
        sheet.append(values)

    for index, column in enumerate(columns, start=1):
        letter = openpyxl.utils.get_column_letter(index)
        sheet.column_dimensions[letter].width = max(12, min(len(column) + 4, 30))
        if column in _DEMO_TEXT_COLUMNS:
            number_format = "@"
        elif column in _DEMO_DATE_COLUMNS:
            number_format = "yyyy-mm-dd"
        else:
            continue
        # Set on the column, so it also governs the rows someone types in
        # underneath the examples -- which is where their real data goes.
        #
        # Deliberately not a loop over the first N rows setting each cell:
        # that materialises N blank rows into the file, and the workbook
        # then opens claiming two thousand rows of nothing.
        sheet.column_dimensions[letter].number_format = number_format
        for row_index in range(2, len(rows) + 2):
            sheet.cell(row=row_index, column=index).number_format = number_format

    sheet.freeze_panes = "A2"

    buf = io.BytesIO()
    workbook.save(buf)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _lookup_by_name(items_by_lower_name, raw_name, label, row_errors):
    """Case-insensitive name lookup for an optional bulk-import column
    (BACE property / festival / seva type / Live To Give purpose). Blank
    input is fine (returns None, no error) -- a name that doesn't match
    anything is treated as a row error rather than silently ignored, so a
    typo doesn't quietly drop data instead of failing loudly."""
    name = (raw_name or "").strip()
    if not name:
        return None
    match = items_by_lower_name.get(name.lower())
    if not match:
        row_errors.append(f"{label} '{name}' not found")
        return None
    return match.id


@bp.route("/donations/bulk-import", methods=["GET", "POST"])
@login_required
@admin_role_required
def bulk_import_donations():
    """Bulk-logs offline donations (cash/cheque/bank transfer) from an
    uploaded CSV -- the same underlying steps as the single-entry "Log
    Offline Donation" form (manual_donation), just looped over many rows.
    Each successful row gets everything an individual manual donation
    gets: a donor record (matched/created via the same find_or_create_donor
    PAN->phone->email dedup used everywhere else), a real receipt number
    from the shared ReceiptCounter sequence, and a generated receipt PDF
    stored on the donation -- so cancellation, exports, statements, and
    Form 10BD all treat imported rows identically to hand-entered ones.

    Rows are validated and processed independently -- one bad row (unknown
    campaign, invalid PAN, malformed date, ...) is skipped with a reason
    rather than failing the whole file. Nothing is committed for a row
    until it passes every check.

    Email/WhatsApp receipt notifications are opt-in per import (default
    off): bulk imports are usually historical/backlog entries, and a donor
    who gave three months ago typically shouldn't get a receipt "just
    arriving" today.
    """
    if request.method == "GET":
        # Single Entry and Bulk Upload live on one merged page/nav tab now
        # (offline_donation.html, served by manual_donation()) -- this
        # route still owns the actual upload processing (POST below), but
        # a direct GET here just lands on that page with Bulk selected.
        return redirect(url_for("admin.manual_donation", tab="bulk"))

    file = request.files.get("csv_file")
    if not file or not file.filename:
        flash("Please choose a CSV file to upload.")
        return redirect(url_for("admin.manual_donation", tab="bulk"))

    send_notifications = request.form.get("send_notifications") == "yes"

    # Dry run: validate the whole file and report what would happen,
    # writing nothing. An import of a few hundred donations issues real
    # receipt numbers from a shared sequence and can email donors -- it is
    # not something to discover the shape of by running it.
    preview = request.form.get("action") == "preview"

    try:
        reader = _table_from_upload(file)
        fieldnames = {(f or "").strip() for f in (reader.fieldnames or [])}
    except _UploadReadError as exc:
        # Says what's actually wrong with the file. The generic message
        # below used to cover this case too, which told someone whose
        # Excel file needed re-saving to upload a CSV instead.
        flash(str(exc))
        return redirect(url_for("admin.manual_donation", tab="bulk"))
    except Exception:
        flash("Couldn't read that file -- please upload a CSV or Excel (.xlsx) file.")
        return redirect(url_for("admin.manual_donation", tab="bulk"))

    missing = [c for c in BULK_IMPORT_REQUIRED_COLUMNS if c not in fieldnames]
    if missing:
        flash(
            "That CSV is missing required column(s): " + ", ".join(missing)
            + ". Download the demo file below for the full column list."
        )
        return redirect(url_for("admin.manual_donation", tab="bulk"))

    campaigns_by_name = {c.name.strip().lower(): c for c in Campaign.query.all()}
    bace_by_name = {b.name.strip().lower(): b for b in BaceProperty.query.all()}
    festivals_by_name = {f.name.strip().lower(): f for f in Festival.query.all()}
    seva_by_name = {s.name.strip().lower(): s for s in SevaType.query.all()}
    purposes_by_name = {p.name.strip().lower(): p for p in LiveToGivePurpose.query.all()}

    results = []
    created = 0
    # Dates where the day-first reading of a D/M/Y value was a guess --
    # reported at the end so the operator can spot-check. See
    # _parse_import_date.
    ambiguous_dates = []

    for line_num, raw_row in enumerate(reader, start=2):  # header is line 1
        row = {(k or "").strip(): (v or "").strip() for k, v in raw_row.items() if k}
        row_errors = []

        full_name = row.get("full_name", "")
        if not full_name:
            row_errors.append("full_name is required")

        campaign_name = row.get("campaign_name", "")
        campaign = campaigns_by_name.get(campaign_name.lower()) if campaign_name else None
        if not campaign_name:
            row_errors.append("campaign_name is required")
        elif not campaign:
            row_errors.append(f"campaign '{campaign_name}' not found")

        amount = None
        amount_raw = row.get("amount", "")
        try:
            amount = float(amount_raw)
            if amount <= 0:
                row_errors.append("amount must be greater than 0")
        except ValueError:
            row_errors.append(f"invalid amount '{amount_raw}'")

        payment_mode = (row.get("payment_mode") or "cash").lower()
        # "online" is accepted here for the same reason it's offered on the
        # single-entry form above it: a payment made online but recorded by
        # hand (collected via a separate gateway, or one this system never
        # saw). Kept in step with that dropdown deliberately -- a mode that
        # can be entered one-at-a-time but rejected in bulk would be a
        # confusing thing to discover halfway through an import.
        if payment_mode not in ("cash", "cheque", "bank_transfer", "online"):
            row_errors.append(
                f"payment_mode must be cash, cheque, bank_transfer, or online (got '{payment_mode}')"
            )

        # Tolerates Excel's locale reformatting -- see _parse_import_date.
        # This importer used to accept YYYY-MM-DD and nothing else, while
        # the donor, camp and legacy importers all took the reformatted
        # form, so the same reviewed-in-Excel file imported through three
        # tabs and failed on this one.
        donation_date = _import_datetime(
            row.get("donation_date"), "donation_date", row_errors, ambiguous=ambiguous_dates)

        pan = row.get("pan", "").upper()
        if pan and not is_valid_pan(pan):
            row_errors.append(f"invalid PAN '{pan}'")
        row["pan"] = pan

        if not is_valid_phone(row.get("phone")):
            row_errors.append(f"invalid phone '{row.get('phone')}' (expected a 10-digit mobile number)")
        if not is_valid_phone(row.get("whatsapp_number")):
            row_errors.append(f"invalid whatsapp_number '{row.get('whatsapp_number')}' (expected a 10-digit mobile number)")

        high_value_error = high_value_pan_address_error(amount, pan, row.get("address"))
        if high_value_error:
            row_errors.append(high_value_error)

        bace_property_id = _lookup_by_name(bace_by_name, row.get("bace_property_name"), "BACE property", row_errors)
        festival_id = _lookup_by_name(festivals_by_name, row.get("festival_name"), "festival", row_errors)
        seva_type_id = _lookup_by_name(seva_by_name, row.get("seva_type_name"), "seva type", row_errors)
        live_to_give_purpose_id = _lookup_by_name(
            purposes_by_name, row.get("live_to_give_purpose_name"), "donation purpose", row_errors
        )

        receipt_type = (row.get("receipt_type") or "").lower()
        if receipt_type == "80g":
            is_80g_requested = True
        elif receipt_type == "non80g":
            is_80g_requested = False
        else:
            is_80g_requested = None

        # _create_offline_donation enforces this, but preview has to know
        # about it too: a preview that promises to import a row the real
        # run will skip is worse than no preview.
        reference_error = _payment_reference_error(
            payment_mode, row.get("cheque_number"), row.get("bank_transaction_id"))
        if reference_error:
            row_errors.append(reference_error)

        if row_errors:
            results.append({"line": line_num, "name": full_name or "(blank)", "ok": False, "errors": row_errors})
            continue

        if preview:
            # Everything above is validation and lookups -- nothing has
            # been written. Record what *would* happen and move on.
            results.append({
                "line": line_num, "name": full_name, "ok": True, "preview": True,
                "amount": amount, "campaign": campaign.name,
                "donation_date": donation_date, "payment_mode": payment_mode,
            })
            created += 1
            continue

        # Same shared path manual_donation() uses -- donor find/create +
        # Donation row + receipt number in one transaction (rolled back
        # whole on any failure, so a bad row never leaves a half-written
        # record), then PDF generation and notification-kickoff each
        # guarded separately. Row-level validation above has already
        # ruled out the common mistakes (bad campaign/amount/date/PAN/
        # phone); this call covers whatever that validation doesn't,
        # same as it does for the single-entry form.
        result = _create_offline_donation(
            donor_data=row,
            campaign=campaign,
            amount=amount,
            payment_mode=payment_mode,
            donation_date=donation_date,
            recorded_by=current_user.username,
            bace_property_id=bace_property_id,
            festival_id=festival_id,
            seva_type_id=seva_type_id,
            live_to_give_purpose_id=live_to_give_purpose_id,
            is_80g_requested=is_80g_requested,
            cheque_number=row.get("cheque_number"),
            cheque_bank_name=row.get("cheque_bank_name"),
            bank_transaction_id=row.get("bank_transaction_id"),
            remarks=row.get("remarks"),
            send_notifications=send_notifications,
        )
        if not result["ok"]:
            results.append({"line": line_num, "name": full_name or "(blank)", "ok": False, "errors": [result["error"]]})
            continue

        row_result = {"line": line_num, "name": full_name, "ok": True, "receipt_number": result["receipt_number"]}
        if not result["pdf_ok"]:
            row_result["pdf_ok"] = False
        results.append(row_result)
        created += 1

    skipped = len(results) - created
    if preview:
        bulk_msg = (
            f"Preview only -- nothing has been saved. {created} row(s) would be "
            f"imported, {skipped} skipped."
        )
    else:
        bulk_msg = f"Bulk import finished: {created} donation(s) created, {skipped} skipped."
    bulk_msg += _ambiguous_date_warning(ambiguous_dates)
    flash(bulk_msg)
    return render_template(
        "admin/offline_donation.html", active_tab="bulk",
        bulk_results=results, bulk_created=created, bulk_skipped=skipped,
        bulk_preview=preview, bulk_summary=_preview_summary(results),
        **_offline_donation_form_context(),
    )


# Same idea as BULK_IMPORT_* above, but for migrating history from before
# this website existed rather than logging new offline donations -- see
# import_legacy_donations() for how the two differ.
LEGACY_IMPORT_REQUIRED_COLUMNS = ["full_name", "campaign_name", "amount", "donation_date"]
LEGACY_IMPORT_COLUMNS = [
    "full_name", "phone", "whatsapp_number", "email", "pan", "address", "city", "state", "pincode",
    "campaign_name", "amount", "payment_mode", "donation_date", "receipt_number", "is_80g_requested",
    "cheque_number", "cheque_bank_name", "bank_transaction_id", "remarks",
]
LEGACY_IMPORT_DEMO_ROWS = [
    {
        "full_name": "Gopal Krishna Das", "phone": "9811122233", "whatsapp_number": "", "email": "gopal@example.com",
        "pan": "ABCDE1234F", "address": "45 Preet Vihar", "city": "Delhi", "state": "Delhi", "pincode": "110092",
        "campaign_name": "Temple Construction", "amount": "11000", "payment_mode": "cash",
        "donation_date": "2023-06-10", "receipt_number": "OLD/2023/00456", "is_80g_requested": "",
        "cheque_number": "", "cheque_bank_name": "", "bank_transaction_id": "", "remarks": "",
    },
    {
        "full_name": "Radha Rani Devi", "phone": "9822233344", "whatsapp_number": "", "email": "",
        "pan": "", "address": "", "city": "", "state": "", "pincode": "",
        "campaign_name": "General Donations", "amount": "2500", "payment_mode": "cheque",
        "donation_date": "2024-01-22", "receipt_number": "OLD/2024/00112", "is_80g_requested": "",
        "cheque_number": "998877", "cheque_bank_name": "SBI", "bank_transaction_id": "", "remarks": "",
    },
    {
        "full_name": "Nitai Chandra", "phone": "", "whatsapp_number": "9933344455", "email": "nitai@example.com",
        "pan": "FGHIJ5678K", "address": "", "city": "", "state": "", "pincode": "",
        "campaign_name": "Annadan", "amount": "7500", "payment_mode": "bank_transfer",
        "donation_date": "2024-11-03", "receipt_number": "", "is_80g_requested": "",
        "cheque_number": "", "cheque_bank_name": "", "bank_transaction_id": "UTR2024110312345", "remarks": "",
    },
    {
        "full_name": "Meera Krishnan", "phone": "9877712345", "whatsapp_number": "", "email": "",
        "pan": "", "address": "", "city": "", "state": "", "pincode": "",
        "campaign_name": "Live To Give", "amount": "1100", "payment_mode": "online",
        "donation_date": "2025-03-14", "receipt_number": "", "is_80g_requested": "no",
        "cheque_number": "", "cheque_bank_name": "", "bank_transaction_id": "pay_ABC123", "remarks": "Food for Life",
    },
]


@bp.route("/donations/import-legacy/demo.csv")
@login_required
@admin_role_required
def import_legacy_demo_csv():
    """Template CSV for migrating pre-website donor/donation history."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=LEGACY_IMPORT_COLUMNS)
    writer.writeheader()
    writer.writerows(LEGACY_IMPORT_DEMO_ROWS)
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=legacy_donations_demo.csv"},
    )


@bp.route("/donations/import-legacy/demo.xlsx")
@login_required
@admin_role_required
def import_legacy_demo_xlsx():
    return _demo_workbook(
        LEGACY_IMPORT_COLUMNS, LEGACY_IMPORT_DEMO_ROWS, "legacy_donations_demo.xlsx")


@bp.route("/donations/import-legacy", methods=["GET", "POST"])
@login_required
@admin_role_required
def import_legacy_donations():
    """Migrates pre-website donor/donation history whose receipts were
    already issued (on paper, or through whatever system you used before)
    -- this is the web-UI equivalent of import_legacy_data.py, for anyone
    who'd rather upload a file than run a script on the server.

    Differs from "Bulk Import" (bulk_import_donations) in the ways that
    matter for already-issued receipts:
      - `receipt_number` is read from the CSV and kept as-is when given,
        instead of being generated from this site's own numbering
        sequence -- your old receipt numbers stay the numbers of record.
        Leave the column blank on a row to auto-generate one from this
        site's sequence instead (e.g. for old records where the original
        number is lost).
      - No PDF is generated for imported rows by default (the real
        receipt already exists on paper/elsewhere) -- ticking "Generate
        PDF receipts" creates one in this site's own layout for every
        row, which will NOT match whatever the original receipt looked
        like. Donations with no PDF here still show correctly in every
        report, statement, and total; the receipt link on this site just
        won't have a file behind it unless you opt in.
      - Email/WhatsApp receipt notifications are never sent -- these
        donors already received their receipt at the time, so importing
        them now should never trigger a fresh notification.

    Donor de-duplication uses the same PAN -> phone -> email matching as
    everywhere else, so importing won't create duplicates of donors who
    already exist from live donations.
    """
    if request.method == "GET":
        return render_template("admin/import_legacy_donations.html", results=None)

    file = request.files.get("csv_file")
    if not file or not file.filename:
        flash("Please choose a CSV file to upload.")
        return redirect(url_for("admin.import_legacy_donations"))

    generate_pdfs = request.form.get("generate_pdfs") == "yes"
    preview = request.form.get("action") == "preview"  # dry run -- see bulk_import_donations

    try:
        reader = _table_from_upload(file)
        fieldnames = {(f or "").strip() for f in (reader.fieldnames or [])}
    except _UploadReadError as exc:
        # Says what's actually wrong with the file. The generic message
        # below used to cover this case too, which told someone whose
        # Excel file needed re-saving to upload a CSV instead.
        flash(str(exc))
        return redirect(url_for("admin.import_legacy_donations"))
    except Exception:
        flash("Couldn't read that file -- please upload a CSV or Excel (.xlsx) file.")
        return redirect(url_for("admin.import_legacy_donations"))

    missing = [c for c in LEGACY_IMPORT_REQUIRED_COLUMNS if c not in fieldnames]
    if missing:
        flash(
            "That CSV is missing required column(s): " + ", ".join(missing)
            + ". Download the demo file below for the full column list."
        )
        return redirect(url_for("admin.import_legacy_donations"))

    campaigns_by_name = {c.name.strip().lower(): c for c in Campaign.query.all()}
    org_cfg = _org_cfg()
    results = []
    created = 0
    no_receipt_pdf_skipped = 0
    ambiguous_dates = []  # see _parse_import_date

    for line_num, raw_row in enumerate(reader, start=2):  # header is line 1
        row = {(k or "").strip(): (v or "").strip() for k, v in raw_row.items() if k}
        row_errors = []

        full_name = row.get("full_name", "")
        if not full_name:
            row_errors.append("full_name is required")

        campaign_name = row.get("campaign_name", "")
        campaign = campaigns_by_name.get(campaign_name.lower()) if campaign_name else None
        if not campaign_name:
            row_errors.append("campaign_name is required")
        elif not campaign:
            row_errors.append(f"campaign '{campaign_name}' not found -- create it first under Admin > Campaigns")

        amount = None
        amount_raw = row.get("amount", "")
        try:
            amount = float(amount_raw)
            if amount <= 0:
                row_errors.append("amount must be greater than 0")
        except ValueError:
            row_errors.append(f"invalid amount '{amount_raw}'")

        payment_mode = (row.get("payment_mode") or "cash").lower()
        if payment_mode not in ("cash", "cheque", "bank_transfer", "online"):
            payment_mode = "cash"  # unrecognised is treated as cash, same as the CLI importer, rather than skipped

        # Was an inline copy of _parse_import_date's fallback. Same rule,
        # one implementation, so the two can't drift.
        donation_date = _import_datetime(
            row.get("donation_date"), "donation_date", row_errors, ambiguous=ambiguous_dates)

        pan = row.get("pan", "").upper()
        if pan and not is_valid_pan(pan):
            row_errors.append(f"invalid PAN '{pan}'")
        row["pan"] = pan

        if not is_valid_phone(row.get("phone")):
            row_errors.append(f"invalid phone '{row.get('phone')}' (expected a 10-digit mobile number)")
        if not is_valid_phone(row.get("whatsapp_number")):
            row_errors.append(f"invalid whatsapp_number '{row.get('whatsapp_number')}' (expected a 10-digit mobile number)")

        # Deliberately NOT enforcing high_value_pan_address_error() here --
        # unlike bulk_import_donations/manual_donation (which issue a real
        # receipt number through this app), legacy-imported rows are
        # historical records for donations whose receipts were already
        # issued under the old system (see existing_receipt below and
        # generate_pdfs handling further down) -- retroactively requiring
        # PAN/address on old external receipts wouldn't fix anything real
        # and would just block digitizing otherwise-valid historical data.

        is_80g_raw = (row.get("is_80g_requested") or "").strip().lower()
        if is_80g_raw in ("yes", "y", "80g", "true", "1"):
            is_80g_requested = True
        elif is_80g_raw in ("no", "n", "non80g", "non-80g", "false", "0"):
            is_80g_requested = False
        elif is_80g_raw:
            row_errors.append(f"is_80g_requested must be yes/no/blank (got '{is_80g_raw}')")
            is_80g_requested = None
        else:
            is_80g_requested = None  # blank -- falls back to the campaign's own default

        existing_receipt = (row.get("receipt_number") or "").strip() or None

        if row_errors:
            results.append({"line": line_num, "name": full_name or "(blank)", "ok": False, "errors": row_errors})
            continue

        if preview:
            results.append({
                "line": line_num, "name": full_name, "ok": True, "preview": True,
                "amount": amount, "campaign": campaign.name,
                "donation_date": donation_date, "payment_mode": payment_mode,
            })
            created += 1
            continue

        try:
            donor = find_or_create_donor(row)

            donation = Donation(
                donor_id=donor.id,
                campaign_id=campaign.id,
                amount=amount,
                payment_mode=payment_mode,
                status="success",
                donation_date=donation_date,
                is_80g_requested=is_80g_requested,
                cheque_number=(row.get("cheque_number") or "")[:50] or None,
                cheque_bank_name=(row.get("cheque_bank_name") or "")[:150] or None,
                bank_transaction_id=(row.get("bank_transaction_id") or "")[:100] or None,
                remarks=(row.get("remarks") or "").strip()[:300] or "Imported from legacy records",
                recorded_by=f"legacy import ({current_user.username})",
            )
            db.session.add(donation)
            db.session.flush()

            # financial_year is always computed from the actual donation
            # date regardless of receipt number -- it's what Form 10BD and
            # every annual report group by.
            donation.financial_year = get_financial_year(donation_date)

            if existing_receipt:
                donation.receipt_number = existing_receipt[:50]
            else:
                # Deliberately NOT auto-generating one from this site's own
                # sequence (032511/ISK500000...) here -- unlike a brand new
                # donation, a legacy row with no receipt_number in the CSV
                # usually means no receipt was ever actually issued for it
                # (or the number just wasn't captured in the export), and
                # minting a fresh number now would misrepresent it as if
                # this site had issued an official receipt at the time.
                # The donation still counts correctly everywhere (totals,
                # Analytics, Form 10BD by financial_year) -- it just has no
                # receipt number on file, same as it had none before.
                donation.receipt_number = None

            if generate_pdfs and donation.receipt_number:
                donation.receipt_pdf = generate_receipt_pdf(donation, donor, campaign, org_cfg)
            elif generate_pdfs:
                no_receipt_pdf_skipped += 1

            db.session.commit()

            results.append({"line": line_num, "name": full_name, "ok": True, "receipt_number": donation.receipt_number})
            created += 1
        except Exception as exc:
            db.session.rollback()
            current_app.logger.exception("Legacy import failed on row %s", line_num)
            reason = "duplicate receipt number" if "unique" in str(exc).lower() else f"unexpected error ({exc})"
            results.append({
                "line": line_num, "name": full_name or "(blank)", "ok": False,
                "errors": [f"row skipped -- {reason}"],
            })

    skipped = len(results) - created
    if preview:
        flash_msg = (
            f"Preview only -- nothing has been saved. {created} row(s) would be "
            f"imported, {skipped} skipped."
        )
    else:
        flash_msg = f"Legacy import finished: {created} donation(s) imported, {skipped} skipped."
    flash_msg += _ambiguous_date_warning(ambiguous_dates)
    if no_receipt_pdf_skipped:
        flash_msg += (
            f" {no_receipt_pdf_skipped} row(s) had no receipt_number, so no PDF was generated for "
            "them even though 'Generate PDF receipts' was ticked -- add a receipt_pdf later from the "
            "donation's own page if one turns up, or leave it as an on-file donation with no receipt."
        )
    flash(flash_msg)
    return render_template(
        "admin/import_legacy_donations.html", results=results, created=created,
        skipped=skipped, preview=preview, summary=_preview_summary(results))


# Donor-only master-data import -- no donation records are created here.
# Distinct from BULK_IMPORT_*/LEGACY_IMPORT_* above (which both import a
# donation transaction per row): this is for uploading/updating the
# contact + relationship + family fields on Donor records directly, e.g.
# a spreadsheet of existing known donors that predates this website.
DONOR_IMPORT_REQUIRED_COLUMNS = ["full_name"]
DONOR_IMPORT_COLUMNS = [
    "full_name", "phone", "whatsapp_number", "email", "pan",
    "address", "city", "state", "pincode",
    "donor_type", "connected_preacher_name", "donation_frequency", "gifts",
    "dob", "father_dob", "mother_dob", "wife_dob", "marriage_anniversary",
    "additional_info",
]
DONOR_IMPORT_DEMO_ROWS = [
    {
        "full_name": "Ramesh Kumar", "phone": "9876543210", "whatsapp_number": "", "email": "ramesh@example.com",
        "pan": "ABCDE1234F", "address": "12 MG Road", "city": "Delhi", "state": "Delhi", "pincode": "110001",
        "donor_type": "iyf", "connected_preacher_name": "", "donation_frequency": "monthly", "gifts": "",
        "dob": "1985-06-12", "father_dob": "", "mother_dob": "", "wife_dob": "", "marriage_anniversary": "",
        "additional_info": "",
    },
    {
        "full_name": "Sita Devi", "phone": "9123456780", "whatsapp_number": "", "email": "",
        "pan": "", "address": "45 Ring Road", "city": "Delhi", "state": "Delhi", "pincode": "110024",
        "donor_type": "live_to_give", "connected_preacher_name": "", "donation_frequency": "quarterly",
        "gifts": "Bhagavad Gita set", "dob": "1978-11-02", "father_dob": "", "mother_dob": "",
        "wife_dob": "", "marriage_anniversary": "2001-02-14", "additional_info": "Prefers WhatsApp contact",
    },
    {
        "full_name": "Amit Sharma", "phone": "", "whatsapp_number": "9988776655", "email": "amit@example.com",
        "pan": "FGHIJ5678K", "address": "", "city": "", "state": "", "pincode": "",
        "donor_type": "iyf", "connected_preacher_name": "", "donation_frequency": "one_time", "gifts": "",
        "dob": "", "father_dob": "", "mother_dob": "", "wife_dob": "", "marriage_anniversary": "",
        "additional_info": "",
    },
]


@bp.route("/donors/import/demo.csv")
@login_required
@admin_role_required
def import_donors_demo_csv():
    """A ready-to-edit template CSV -- every column the donor importer
    understands, header row plus a few example rows. Preacher names are
    left blank in the demo so it imports cleanly out of the box even
    before any Preachers have been added under Admin -> Preachers."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=DONOR_IMPORT_COLUMNS)
    writer.writeheader()
    writer.writerows(DONOR_IMPORT_DEMO_ROWS)
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=donor_data_demo.csv"},
    )


@bp.route("/donors/import/demo.xlsx")
@login_required
@admin_role_required
def import_donors_demo_xlsx():
    return _demo_workbook(
        DONOR_IMPORT_COLUMNS, DONOR_IMPORT_DEMO_ROWS, "donor_data_demo.xlsx")


def _parse_import_date(raw, label, row_errors, required=False, ambiguous=None):
    """The one date parser for every CSV importer in this file.

    YYYY-MM-DD is the canonical format (what every demo template uses),
    but a spreadsheet that's been opened and re-saved in Excel or Google
    Sheets for a quick review commonly gets its date cells silently
    reformatted to a locale style first -- 2024-01-22 comes back as
    22/01/2024 or 22/01/24. Nobody notices until the import rejects half
    the file, and the natural next move (retype every date) is worse than
    the problem. So day-first D/M/Y with a 2- or 4-digit year is accepted
    as a fallback.

    `required=True` makes a blank value an error, for the importers where
    the date is a required column. Left False, blank means "leave whatever
    is already on file untouched", which is what the donor importer wants.

    `ambiguous`, if given a list, collects values where the day-first
    reading was a *guess* -- both numbers 12 or under, so 01/08/2026 could
    equally be 1 August or 8 January. Day-first is right for a file
    produced in an Indian locale and wrong for a US one, and on a donation
    the difference can land the receipt in the wrong financial year, which
    then flows into Form 10BD. The caller surfaces the count so the
    operator can spot-check rather than find out at audit.
    """
    raw = (raw or "").strip()
    if not raw:
        if required:
            row_errors.append(f"{label} is required")
        return None

    try:
        return datetime.datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError:
        pass

    # Year-first with other separators (2026/08/01) -- some Google Sheets
    # locales write this. Unambiguous, since a 4-digit leading number can
    # only be the year, so it needs no day-first guess.
    m = re.match(r"^(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})$", raw)
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass

    m = re.match(r"^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})$", raw)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if year < 100:
            year += 2000 if year < 30 else 1900
        try:
            parsed = datetime.date(year, month, day)
        except ValueError:
            parsed = None
        if parsed:
            if ambiguous is not None and day <= 12 and month <= 12:
                ambiguous.append(raw)
            return parsed

    row_errors.append(f"invalid {label} '{raw}' (expected YYYY-MM-DD)")
    return None


def _preview_summary(results):
    """Totals for a dry run: what the file adds up to, so the numbers can
    be checked against the cash book or the bank before anything is
    written. A row count alone doesn't catch an amount column that landed
    a factor of ten out."""
    rows = [r for r in results if r.get("ok") and r.get("preview")]
    if not rows:
        return None
    dates = [r["donation_date"] for r in rows if r.get("donation_date")]
    return {
        "count": len(rows),
        "total": sum(r.get("amount") or 0 for r in rows),
        "first_date": min(dates) if dates else None,
        "last_date": max(dates) if dates else None,
        "campaigns": sorted({r["campaign"] for r in rows if r.get("campaign")}),
    }


def _ambiguous_date_warning(ambiguous):
    """Message appended to an import's summary when day-first D/M/Y dates
    were guessed at. Silent when there's nothing to say."""
    if not ambiguous:
        return ""
    sample = ", ".join(sorted(set(ambiguous))[:3])
    return (
        f" Note: {len(ambiguous)} date(s) were written as D/M/Y and read day-first "
        f"(e.g. {sample}). That's correct for a file saved in an Indian locale. If it came "
        "from a US-locale spreadsheet the day and month are swapped -- check those donations "
        "before issuing receipts, since the date decides the financial year."
    )


def _import_datetime(raw, label, row_errors, ambiguous=None):
    """_parse_import_date for the two donation importers, which store a
    datetime and always require the date."""
    parsed = _parse_import_date(raw, label, row_errors, required=True, ambiguous=ambiguous)
    if parsed is None:
        return None
    return datetime.datetime.combine(parsed, datetime.time())


def _resolve_preacher_import(preachers_by_name, raw_name, row_errors):
    """Returns (should_update, preacher_id_or_None) for the
    connected_preacher_name import column. Blank means 'leave the donor's
    existing preacher assignment untouched'; 'none' / 'not assigned' /
    'unassigned' explicitly clears it; anything else must case-insensitively
    match an existing Preacher's name (manage those under Admin -> Preachers
    first) or the row fails."""
    name = (raw_name or "").strip()
    if not name:
        return False, None
    if name.lower() in ("none", "not assigned", "unassigned"):
        return True, None
    match = preachers_by_name.get(name.lower())
    if not match:
        row_errors.append(f"preacher '{name}' not found")
        return False, None
    return True, match.id


@bp.route("/donors/import", methods=["GET", "POST"])
@login_required
@admin_role_required
def import_donors():
    """Bulk-imports/updates donor master data -- contact details plus the
    Donor Type / Connected Preacher / Donation Frequency / family-date
    fields normally filled in one-by-one under Admin -> Donors -> Edit --
    from a CSV. No donation records are created here.

    Existing donors are matched using the same PAN -> phone+name -> email+name
    logic as every other donor-touching form (find_or_create_donor from
    public.py) -- a PAN match is always treated as the same person, while a
    phone/email match additionally requires the name on the row to agree
    with the name already on file (a shared family phone number shouldn't
    let one row overwrite a different family member's details). A genuine
    match updates that donor's fields following the "new value wins, blank
    leaves the existing value alone" convention -- so re-uploading the same
    file, or a file that only has a few columns filled in for some rows,
    never wipes out data that's already on file.
    """
    if request.method == "GET":
        return render_template("admin/import_donors.html", results=None)

    file = request.files.get("csv_file")
    if not file or not file.filename:
        flash("Please choose a CSV file to upload.")
        return redirect(url_for("admin.import_donors"))

    try:
        reader = _table_from_upload(file)
        fieldnames = {(f or "").strip() for f in (reader.fieldnames or [])}
    except _UploadReadError as exc:
        # Says what's actually wrong with the file. The generic message
        # below used to cover this case too, which told someone whose
        # Excel file needed re-saving to upload a CSV instead.
        flash(str(exc))
        return redirect(url_for("admin.import_donors"))
    except Exception:
        flash("Couldn't read that file -- please upload a CSV or Excel (.xlsx) file.")
        return redirect(url_for("admin.import_donors"))

    missing = [c for c in DONOR_IMPORT_REQUIRED_COLUMNS if c not in fieldnames]
    if missing:
        flash(
            "That CSV is missing required column(s): " + ", ".join(missing)
            + ". Download the demo file below for the full column list."
        )
        return redirect(url_for("admin.import_donors"))

    preachers_by_name = {p.name.strip().lower(): p for p in Preacher.query.all()}

    preview = request.form.get("action") == "preview"  # dry run -- see bulk_import_donations
    results = []
    created = 0
    updated = 0

    for line_num, raw_row in enumerate(reader, start=2):  # header is line 1
        row = {(k or "").strip(): (v or "").strip() for k, v in raw_row.items() if k}
        row_errors = []

        full_name = row.get("full_name", "")
        if not full_name:
            row_errors.append("full_name is required")

        pan = row.get("pan", "").upper()
        if pan and not is_valid_pan(pan):
            row_errors.append(f"invalid PAN '{pan}'")
        row["pan"] = pan

        if not is_valid_phone(row.get("phone")):
            row_errors.append(f"invalid phone '{row.get('phone')}' (expected a 10-digit mobile number)")
        if not is_valid_phone(row.get("whatsapp_number")):
            row_errors.append(f"invalid whatsapp_number '{row.get('whatsapp_number')}' (expected a 10-digit mobile number)")

        donor_type_raw = row.get("donor_type", "").lower()
        if donor_type_raw and donor_type_raw not in DONOR_TYPES:
            row_errors.append(f"donor_type must be one of {', '.join(DONOR_TYPES)} (got '{donor_type_raw}')")

        frequency_raw = row.get("donation_frequency", "").lower()
        if frequency_raw and frequency_raw not in DONATION_FREQUENCIES:
            row_errors.append(
                f"donation_frequency must be one of {', '.join(DONATION_FREQUENCIES)} (got '{frequency_raw}')"
            )

        update_preacher, preacher_id = _resolve_preacher_import(
            preachers_by_name, row.get("connected_preacher_name"), row_errors
        )

        dob = _parse_import_date(row.get("dob"), "dob", row_errors)
        father_dob = _parse_import_date(row.get("father_dob"), "father_dob", row_errors)
        mother_dob = _parse_import_date(row.get("mother_dob"), "mother_dob", row_errors)
        wife_dob = _parse_import_date(row.get("wife_dob"), "wife_dob", row_errors)
        marriage_anniversary = _parse_import_date(row.get("marriage_anniversary"), "marriage_anniversary", row_errors)

        if row_errors:
            results.append({"line": line_num, "name": full_name or "(blank)", "ok": False, "errors": row_errors})
            continue

        try:
            # Mirrors find_or_create_donor's own PAN -> phone -> email
            # matching (including normalize_phone()) so we know up front
            # whether this row will create a new donor or update an
            # existing one (for the results table).
            phone_v = normalize_phone(row.get("phone"))
            email_v = row.get("email", "").strip().lower()
            existing = None
            if pan:
                existing = Donor.query.filter_by(pan=pan).first()
            if existing is None and phone_v:
                existing = Donor.query.filter_by(phone=phone_v).first()
            if existing is None and email_v:
                existing = Donor.query.filter_by(email=email_v).first()
            was_new = existing is None

            if preview:
                # The matching above is read-only, so the preview can say
                # which rows land on an existing donor -- the thing worth
                # knowing before an import that overwrites donor details.
                results.append({
                    "line": line_num, "name": full_name, "ok": True, "preview": True,
                    "was_new": was_new,
                })
                if was_new:
                    created += 1
                else:
                    updated += 1
                continue

            donor = find_or_create_donor(row)

            if donor_type_raw:
                donor.donor_type = donor_type_raw
            if update_preacher:
                donor.connected_preacher_id = preacher_id
            if frequency_raw:
                donor.donation_frequency = frequency_raw
            gifts_raw = row.get("gifts", "")
            if gifts_raw:
                donor.gifts = gifts_raw[:500]
            if dob:
                donor.dob = dob
            if father_dob:
                donor.father_dob = father_dob
            if mother_dob:
                donor.mother_dob = mother_dob
            if wife_dob:
                donor.wife_dob = wife_dob
            if marriage_anniversary:
                donor.marriage_anniversary = marriage_anniversary
            additional_info_raw = row.get("additional_info", "")
            if additional_info_raw:
                donor.additional_info = additional_info_raw

            db.session.commit()
            results.append({
                "line": line_num, "name": full_name, "ok": True,
                "action": "created" if was_new else "updated", "donor_id": donor.id,
            })
            if was_new:
                created += 1
            else:
                updated += 1
        except Exception as exc:
            db.session.rollback()
            current_app.logger.exception("Donor import failed on row %s", line_num)
            results.append({
                "line": line_num, "name": full_name or "(blank)", "ok": False,
                "errors": [f"unexpected error -- row skipped ({exc})"],
            })

    skipped = len(results) - created - updated
    if preview:
        flash(
            f"Preview only -- nothing has been saved. {created} donor(s) would be "
            f"created and {updated} updated, {skipped} skipped."
        )
    else:
        flash(f"Donor import finished: {created} created, {updated} updated, {skipped} skipped.")
    return render_template(
        "admin/import_donors.html", results=results, created=created, updated=updated,
        skipped=skipped, preview=preview
    )


@bp.route("/campaigns", methods=["GET", "POST"])
@login_required
def campaigns():
    if request.method == "POST":
        if current_user.role != "admin":
            flash("That action requires an administrator account.")
            return redirect(url_for("admin.campaigns"))
        form = request.form
        campaign = Campaign(
            name=form["name"].strip(),
            is_80g=(form.get("is_80g") == "on"),
            description=form.get("description", "").strip() or None,
            target_amount=float(form["target_amount"]) if form.get("target_amount") else None,
            min_amount=float(form["min_amount"]) if form.get("min_amount") else None,
        )
        db.session.add(campaign)
        db.session.flush()
        log_activity("campaign_create", target_type="campaign", target_id=campaign.id, details=f"Created campaign '{campaign.name}'")
        db.session.commit()
        flash(f"Campaign '{campaign.name}' created.")
        return redirect(url_for("admin.campaigns"))

    campaign_list = Campaign.query.order_by(Campaign.is_80g.desc(), Campaign.name).all()
    return render_template("admin/campaigns.html", campaigns=campaign_list)


@bp.route("/campaigns/<int:campaign_id>/toggle", methods=["POST"])
@login_required
@admin_role_required
def toggle_campaign(campaign_id):
    campaign = Campaign.query.get_or_404(campaign_id)
    campaign.is_active = not campaign.is_active
    log_activity(
        "campaign_toggle", target_type="campaign", target_id=campaign.id,
        details=f"'{campaign.name}' set to {'active' if campaign.is_active else 'inactive'}",
    )
    db.session.commit()
    return redirect(url_for("admin.campaigns"))


@bp.route("/campaigns/<int:campaign_id>/edit", methods=["GET", "POST"])
@login_required
@admin_role_required
def campaign_edit(campaign_id):
    campaign = Campaign.query.get_or_404(campaign_id)
    if request.method == "POST":
        form = request.form
        new_name = form.get("name", "").strip()
        if not new_name:
            flash("Campaign name can't be blank.")
            return redirect(url_for("admin.campaign_edit", campaign_id=campaign_id))

        existing = Campaign.query.filter(Campaign.name == new_name, Campaign.id != campaign.id).first()
        if existing:
            flash(f"Another campaign is already named '{new_name}'.")
            return redirect(url_for("admin.campaign_edit", campaign_id=campaign_id))

        campaign.name = new_name
        campaign.is_80g = form.get("is_80g") == "on"
        campaign.description = form.get("description", "").strip() or None
        campaign.target_amount = float(form["target_amount"]) if form.get("target_amount") else None
        campaign.min_amount = float(form["min_amount"]) if form.get("min_amount") else None
        log_activity("campaign_edit", target_type="campaign", target_id=campaign.id, details=f"Edited campaign '{campaign.name}'")
        db.session.commit()
        flash(f"Campaign '{campaign.name}' updated.")
        return redirect(url_for("admin.campaigns"))

    return render_template("admin/campaign_edit.html", campaign=campaign)


@bp.route("/campaigns/<int:campaign_id>/delete", methods=["POST"])
@login_required
@admin_role_required
def campaign_delete(campaign_id):
    campaign = Campaign.query.get_or_404(campaign_id)
    has_donations = Donation.query.filter_by(campaign_id=campaign.id).first() is not None
    if has_donations:
        flash(
            f"Can't delete '{campaign.name}' -- it has donations recorded against it. "
            "Deactivate it instead to hide it from the donation form."
        )
        return redirect(url_for("admin.campaigns"))

    deleted_id, deleted_name = campaign.id, campaign.name
    db.session.delete(campaign)
    log_activity("campaign_delete", target_type="campaign", target_id=deleted_id, details=f"Deleted campaign '{deleted_name}'")
    db.session.commit()
    flash(f"Campaign '{deleted_name}' deleted.")
    return redirect(url_for("admin.campaigns"))


@bp.route("/bace-properties", methods=["GET", "POST"])
@login_required
def bace_properties():
    if request.method == "POST":
        if current_user.role != "admin":
            flash("That action requires an administrator account.")
            return redirect(url_for("admin.bace_properties"))
        name = request.form.get("name", "").strip()
        if not name:
            flash("Property name can't be blank.")
            return redirect(url_for("admin.bace_properties"))
        if BaceProperty.query.filter_by(name=name).first():
            flash(f"A BACE property named '{name}' already exists.")
            return redirect(url_for("admin.bace_properties"))
        db.session.add(BaceProperty(name=name))
        db.session.commit()
        flash(f"BACE property '{name}' added.")
        return redirect(url_for("admin.bace_properties"))

    properties = BaceProperty.query.order_by(BaceProperty.name).all()
    return render_template("admin/bace_properties.html", properties=properties)


@bp.route("/bace-properties/<int:property_id>/toggle", methods=["POST"])
@login_required
@admin_role_required
def toggle_bace_property(property_id):
    prop = BaceProperty.query.get_or_404(property_id)
    prop.is_active = not prop.is_active
    db.session.commit()
    return redirect(url_for("admin.bace_properties"))


@bp.route("/bace-properties/<int:property_id>/edit", methods=["GET", "POST"])
@login_required
@admin_role_required
def bace_property_edit(property_id):
    prop = BaceProperty.query.get_or_404(property_id)
    if request.method == "POST":
        new_name = request.form.get("name", "").strip()
        if not new_name:
            flash("Property name can't be blank.")
            return redirect(url_for("admin.bace_property_edit", property_id=property_id))

        existing = BaceProperty.query.filter(
            BaceProperty.name == new_name, BaceProperty.id != prop.id
        ).first()
        if existing:
            flash(f"Another BACE property is already named '{new_name}'.")
            return redirect(url_for("admin.bace_property_edit", property_id=property_id))

        prop.name = new_name
        db.session.commit()
        flash(f"BACE property renamed to '{prop.name}'.")
        return redirect(url_for("admin.bace_properties"))

    return render_template("admin/bace_property_edit.html", property=prop)


@bp.route("/bace-properties/<int:property_id>/delete", methods=["POST"])
@login_required
@admin_role_required
def bace_property_delete(property_id):
    prop = BaceProperty.query.get_or_404(property_id)
    has_donations = Donation.query.filter_by(bace_property_id=prop.id).first() is not None
    if has_donations:
        flash(
            f"Can't delete '{prop.name}' -- it has donations recorded against it. "
            "Deactivate it instead to hide it from the BACE Contribution form."
        )
        return redirect(url_for("admin.bace_properties"))

    db.session.delete(prop)
    db.session.commit()
    flash(f"BACE property '{prop.name}' deleted.")
    return redirect(url_for("admin.bace_properties"))


def _bace_campaign_or_none():
    """The single "BACE Contribution" campaign that public.bace_rent_form()
    fixes every BACE donation to -- see that route for the same lookup.
    None if the campaign hasn't been set up yet, in which case the log
    page below just shows "no campaign configured" instead of erroring."""
    return Campaign.query.filter_by(name="BACE Contribution").first()


def _apply_bace_contribution_filters(query):
    """Same shape as _apply_donations_filters, but scoped to the BACE
    Contribution campaign and with an extra bace_property_id filter so
    staff can pull up "which BACE has made contribution" for one specific
    property. Shared by the log page and its CSV export."""
    status = request.args.get("status", "success")
    if status != "all":
        query = query.filter_by(status=status)

    bace_property_id = request.args.get("bace_property_id", type=int)
    if bace_property_id:
        query = query.filter_by(bace_property_id=bace_property_id)

    date_from_raw = request.args.get("date_from") or ""
    date_to_raw = request.args.get("date_to") or ""
    try:
        if date_from_raw:
            date_from = datetime.datetime.strptime(date_from_raw, "%Y-%m-%d")
            query = query.filter(Donation.donation_date >= date_from)
        if date_to_raw:
            date_to = datetime.datetime.strptime(date_to_raw, "%Y-%m-%d")
            query = query.filter(Donation.donation_date < date_to + datetime.timedelta(days=1))
    except ValueError:
        date_from_raw = date_to_raw = ""

    return query, {
        "status": status, "bace_property_id": bace_property_id,
        "date_from": date_from_raw, "date_to": date_to_raw,
    }


@bp.route("/bace-contributions")
@login_required
def bace_contributions():
    """Dedicated log for BACE Contribution donations -- unlike the general
    Donations Log (which only shows the parent campaign name), this shows
    which specific BACE property each contribution was for, plus a
    per-property running total so "which BACE has made contribution" is
    answered at a glance instead of needing to open each row."""
    campaign = _bace_campaign_or_none()
    properties = BaceProperty.query.order_by(BaceProperty.name).all()

    if campaign is None:
        return render_template(
            "admin/bace_contributions.html", campaign=None, properties=properties,
            donations=[], pagination=None, summary=[], status="success", bace_property_id=None,
            date_from="", date_to="",
        )

    base_query = Donation.query.filter_by(campaign_id=campaign.id)

    # Per-property totals -- always computed off successful donations only
    # (a failed/pending/cancelled row never actually collected anything),
    # independent of whatever status filter is applied to the table below.
    summary_rows = (
        db.session.query(
            BaceProperty.id, BaceProperty.name, BaceProperty.is_active,
            func.coalesce(func.sum(Donation.amount), 0),
            func.count(Donation.id),
        )
        .outerjoin(Donation, (Donation.bace_property_id == BaceProperty.id) & (Donation.status == "success"))
        .group_by(BaceProperty.id, BaceProperty.name, BaceProperty.is_active)
        .order_by(BaceProperty.name)
        .all()
    )
    summary = [
        {"id": pid, "name": name, "is_active": is_active, "total": total, "count": count}
        for pid, name, is_active, total, count in summary_rows
    ]

    query, filters = _apply_bace_contribution_filters(base_query)
    query = query.order_by(Donation.donation_date.desc())
    page = request.args.get("page", 1, type=int)
    pagination = db.paginate(query, page=page, per_page=DONATIONS_PER_PAGE, error_out=False)

    return render_template(
        "admin/bace_contributions.html", campaign=campaign, properties=properties,
        donations=pagination.items, pagination=pagination, summary=summary, **filters,
    )


@bp.route("/bace-contributions/export")
@login_required
def export_bace_contributions():
    """CSV export of the BACE Contribution log, honoring whatever
    property/status/date filters are currently applied -- same convention
    as export_donations()."""
    campaign = _bace_campaign_or_none()
    if campaign is None:
        flash("The BACE Contribution campaign isn't set up yet.")
        return redirect(url_for("admin.bace_properties"))

    query, _filters = _apply_bace_contribution_filters(Donation.query.filter_by(campaign_id=campaign.id))
    rows = query.order_by(Donation.donation_date.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Receipt No", "Date", "Status", "BACE Property", "Donor Name", "Phone", "Email", "PAN", "Address",
        "Amount", "Payment Mode", "Payment ID", "Order ID", "Recorded By", "Remarks",
    ])
    for d in rows:
        donor = d.donor
        date_str = (
            to_ist(d.donation_date).strftime("%d-%m-%Y %H:%M")
            if d.payment_mode == "online"
            else d.donation_date.strftime("%d-%m-%Y")
        )
        writer.writerow(csv_safe_row([
            d.receipt_number or "",
            date_str,
            d.status,
            d.bace_property.name if d.bace_property else "",
            donor.full_name,
            donor.phone or "",
            donor.email or "",
            donor.pan or "",
            donor.address or "",
            float(d.amount),
            d.payment_mode,
            d.razorpay_payment_id or "",
            d.razorpay_order_id or "",
            d.recorded_by or "",
            d.remarks or "",
        ]))

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=BACE_Contributions.csv"},
    )


@bp.route("/festivals", methods=["GET", "POST"])
@login_required
def festivals():
    if request.method == "POST":
        if current_user.role != "admin":
            flash("That action requires an administrator account.")
            return redirect(url_for("admin.festivals"))
        name = request.form.get("name", "").strip()
        if not name:
            flash("Festival name can't be blank.")
            return redirect(url_for("admin.festivals"))
        if Festival.query.filter_by(name=name).first():
            flash(f"A festival named '{name}' already exists.")
            return redirect(url_for("admin.festivals"))
        event_date_str = request.form.get("event_date")
        try:
            event_date = datetime.datetime.strptime(event_date_str, "%Y-%m-%d").date() if event_date_str else None
        except ValueError:
            flash("That date doesn't look right.")
            return redirect(url_for("admin.festivals"))
        db.session.add(Festival(name=name, event_date=event_date))
        db.session.commit()
        flash(f"Festival '{name}' added.")
        return redirect(url_for("admin.festivals"))

    festival_list = Festival.query.order_by(Festival.event_date.is_(None), Festival.event_date, Festival.name).all()
    return render_template("admin/festivals.html", festivals=festival_list)


@bp.route("/festivals/<int:festival_id>/toggle", methods=["POST"])
@login_required
@admin_role_required
def toggle_festival(festival_id):
    festival = Festival.query.get_or_404(festival_id)
    festival.is_active = not festival.is_active
    db.session.commit()
    return redirect(url_for("admin.festivals"))


@bp.route("/festivals/<int:festival_id>/edit", methods=["GET", "POST"])
@login_required
@admin_role_required
def festival_edit(festival_id):
    festival = Festival.query.get_or_404(festival_id)
    if request.method == "POST":
        new_name = request.form.get("name", "").strip()
        if not new_name:
            flash("Festival name can't be blank.")
            return redirect(url_for("admin.festival_edit", festival_id=festival_id))
        existing = Festival.query.filter(Festival.name == new_name, Festival.id != festival.id).first()
        if existing:
            flash(f"Another festival is already named '{new_name}'.")
            return redirect(url_for("admin.festival_edit", festival_id=festival_id))

        event_date_str = request.form.get("event_date")
        try:
            event_date = datetime.datetime.strptime(event_date_str, "%Y-%m-%d").date() if event_date_str else None
        except ValueError:
            flash("That date doesn't look right.")
            return redirect(url_for("admin.festival_edit", festival_id=festival_id))

        festival.name = new_name
        festival.event_date = event_date
        db.session.commit()
        flash(f"Festival '{festival.name}' updated.")
        return redirect(url_for("admin.festivals"))

    return render_template("admin/festival_edit.html", festival=festival)


@bp.route("/festivals/<int:festival_id>/delete", methods=["POST"])
@login_required
@admin_role_required
def festival_delete(festival_id):
    festival = Festival.query.get_or_404(festival_id)
    has_donations = Donation.query.filter_by(festival_id=festival.id).first() is not None
    if has_donations:
        flash(
            f"Can't delete '{festival.name}' -- it has donations recorded against it. "
            "Deactivate it instead to hide it from the Festival Seva form."
        )
        return redirect(url_for("admin.festivals"))

    db.session.delete(festival)
    db.session.commit()
    flash(f"Festival '{festival.name}' deleted.")
    return redirect(url_for("admin.festivals"))


@bp.route("/seva-types", methods=["GET", "POST"])
@login_required
def seva_types():
    if request.method == "POST":
        if current_user.role != "admin":
            flash("That action requires an administrator account.")
            return redirect(url_for("admin.seva_types"))
        name = request.form.get("name", "").strip()
        if not name:
            flash("Seva type name can't be blank.")
            return redirect(url_for("admin.seva_types"))
        if SevaType.query.filter_by(name=name).first():
            flash(f"A seva type named '{name}' already exists.")
            return redirect(url_for("admin.seva_types"))
        suggested_amount = float(request.form["suggested_amount"]) if request.form.get("suggested_amount") else None
        db.session.add(SevaType(name=name, suggested_amount=suggested_amount))
        db.session.commit()
        flash(f"Seva type '{name}' added.")
        return redirect(url_for("admin.seva_types"))

    seva_type_list = SevaType.query.order_by(SevaType.name).all()
    return render_template("admin/seva_types.html", seva_types=seva_type_list)


@bp.route("/seva-types/<int:seva_type_id>/toggle", methods=["POST"])
@login_required
@admin_role_required
def toggle_seva_type(seva_type_id):
    seva_type = SevaType.query.get_or_404(seva_type_id)
    seva_type.is_active = not seva_type.is_active
    db.session.commit()
    return redirect(url_for("admin.seva_types"))


@bp.route("/seva-types/<int:seva_type_id>/edit", methods=["GET", "POST"])
@login_required
@admin_role_required
def seva_type_edit(seva_type_id):
    seva_type = SevaType.query.get_or_404(seva_type_id)
    if request.method == "POST":
        new_name = request.form.get("name", "").strip()
        if not new_name:
            flash("Seva type name can't be blank.")
            return redirect(url_for("admin.seva_type_edit", seva_type_id=seva_type_id))
        existing = SevaType.query.filter(SevaType.name == new_name, SevaType.id != seva_type.id).first()
        if existing:
            flash(f"Another seva type is already named '{new_name}'.")
            return redirect(url_for("admin.seva_type_edit", seva_type_id=seva_type_id))

        seva_type.name = new_name
        seva_type.suggested_amount = (
            float(request.form["suggested_amount"]) if request.form.get("suggested_amount") else None
        )
        db.session.commit()
        flash(f"Seva type '{seva_type.name}' updated.")
        return redirect(url_for("admin.seva_types"))

    return render_template("admin/seva_type_edit.html", seva_type=seva_type)


@bp.route("/seva-types/<int:seva_type_id>/delete", methods=["POST"])
@login_required
@admin_role_required
def seva_type_delete(seva_type_id):
    seva_type = SevaType.query.get_or_404(seva_type_id)
    has_donations = Donation.query.filter_by(seva_type_id=seva_type.id).first() is not None
    if has_donations:
        flash(
            f"Can't delete '{seva_type.name}' -- it has donations recorded against it. "
            "Deactivate it instead to hide it from the Festival Seva form."
        )
        return redirect(url_for("admin.seva_types"))

    db.session.delete(seva_type)
    db.session.commit()
    flash(f"Seva type '{seva_type.name}' deleted.")
    return redirect(url_for("admin.seva_types"))


@bp.route("/live-to-give-purposes", methods=["GET", "POST"])
@login_required
def live_to_give_purposes():
    if request.method == "POST":
        if current_user.role != "admin":
            flash("That action requires an administrator account.")
            return redirect(url_for("admin.live_to_give_purposes"))
        name = request.form.get("name", "").strip()
        if not name:
            flash("Purpose name can't be blank.")
            return redirect(url_for("admin.live_to_give_purposes"))
        if LiveToGivePurpose.query.filter_by(name=name).first():
            flash(f"A donation purpose named '{name}' already exists.")
            return redirect(url_for("admin.live_to_give_purposes"))
        is_80g = request.form.get("is_80g") == "on"
        db.session.add(LiveToGivePurpose(name=name, is_80g=is_80g))
        db.session.commit()
        flash(f"Donation purpose '{name}' added.")
        return redirect(url_for("admin.live_to_give_purposes"))

    purpose_list = LiveToGivePurpose.query.order_by(LiveToGivePurpose.name).all()
    return render_template("admin/live_to_give_purposes.html", purposes=purpose_list)


@bp.route("/live-to-give-purposes/<int:purpose_id>/toggle", methods=["POST"])
@login_required
@admin_role_required
def toggle_live_to_give_purpose(purpose_id):
    purpose = LiveToGivePurpose.query.get_or_404(purpose_id)
    purpose.is_active = not purpose.is_active
    db.session.commit()
    return redirect(url_for("admin.live_to_give_purposes"))


@bp.route("/live-to-give-purposes/<int:purpose_id>/toggle-80g", methods=["POST"])
@login_required
@admin_role_required
def toggle_live_to_give_purpose_80g(purpose_id):
    """Flips whether this purpose is 80G-eligible -- see
    LiveToGivePurpose.is_80g and Donation.effective_is_80g for what this
    actually controls. Logged since it directly affects which future
    receipts are legally issuable as 80G."""
    purpose = LiveToGivePurpose.query.get_or_404(purpose_id)
    purpose.is_80g = not purpose.is_80g
    log_activity(
        "live_to_give_purpose_80g_toggle", target_type="live_to_give_purpose", target_id=purpose.id,
        details=f"'{purpose.name}' set to {'80G-eligible' if purpose.is_80g else 'Non-80G'}",
    )
    db.session.commit()
    flash(f"'{purpose.name}' is now {'80G-eligible' if purpose.is_80g else 'Non-80G'}.")
    return redirect(url_for("admin.live_to_give_purposes"))


@bp.route("/live-to-give-purposes/<int:purpose_id>/edit", methods=["GET", "POST"])
@login_required
@admin_role_required
def live_to_give_purpose_edit(purpose_id):
    purpose = LiveToGivePurpose.query.get_or_404(purpose_id)
    if request.method == "POST":
        new_name = request.form.get("name", "").strip()
        if not new_name:
            flash("Purpose name can't be blank.")
            return redirect(url_for("admin.live_to_give_purpose_edit", purpose_id=purpose_id))
        existing = LiveToGivePurpose.query.filter(
            LiveToGivePurpose.name == new_name, LiveToGivePurpose.id != purpose.id
        ).first()
        if existing:
            flash(f"Another donation purpose is already named '{new_name}'.")
            return redirect(url_for("admin.live_to_give_purpose_edit", purpose_id=purpose_id))

        purpose.name = new_name
        purpose.is_80g = request.form.get("is_80g") == "on"
        db.session.commit()
        flash(f"Donation purpose renamed to '{purpose.name}'.")
        return redirect(url_for("admin.live_to_give_purposes"))

    return render_template("admin/live_to_give_purpose_edit.html", purpose=purpose)


@bp.route("/live-to-give-purposes/<int:purpose_id>/delete", methods=["POST"])
@login_required
@admin_role_required
def live_to_give_purpose_delete(purpose_id):
    purpose = LiveToGivePurpose.query.get_or_404(purpose_id)
    has_donations = Donation.query.filter_by(live_to_give_purpose_id=purpose.id).first() is not None
    if has_donations:
        flash(
            f"Can't delete '{purpose.name}' -- it has donations recorded against it. "
            "Deactivate it instead to hide it from the Live To Give form."
        )
        return redirect(url_for("admin.live_to_give_purposes"))

    db.session.delete(purpose)
    db.session.commit()
    flash(f"Donation purpose '{purpose.name}' deleted.")
    return redirect(url_for("admin.live_to_give_purposes"))


@bp.route("/preachers", methods=["GET", "POST"])
@login_required
def preachers():
    if request.method == "POST":
        if current_user.role != "admin":
            flash("That action requires an administrator account.")
            return redirect(url_for("admin.preachers"))
        name = request.form.get("name", "").strip()
        if not name:
            flash("Preacher name can't be blank.")
            return redirect(url_for("admin.preachers"))
        if Preacher.query.filter_by(name=name).first():
            flash(f"A preacher named '{name}' already exists.")
            return redirect(url_for("admin.preachers"))
        db.session.add(Preacher(name=name))
        db.session.commit()
        flash(f"Preacher '{name}' added.")
        return redirect(url_for("admin.preachers"))

    preacher_list = Preacher.query.order_by(Preacher.name).all()
    return render_template("admin/preachers.html", preachers=preacher_list)


@bp.route("/preachers/<int:preacher_id>/toggle", methods=["POST"])
@login_required
@admin_role_required
def toggle_preacher(preacher_id):
    preacher = Preacher.query.get_or_404(preacher_id)
    preacher.is_active = not preacher.is_active
    db.session.commit()
    return redirect(url_for("admin.preachers"))


@bp.route("/preachers/<int:preacher_id>/edit", methods=["GET", "POST"])
@login_required
@admin_role_required
def preacher_edit(preacher_id):
    preacher = Preacher.query.get_or_404(preacher_id)
    if request.method == "POST":
        new_name = request.form.get("name", "").strip()
        if not new_name:
            flash("Preacher name can't be blank.")
            return redirect(url_for("admin.preacher_edit", preacher_id=preacher_id))
        existing = Preacher.query.filter(Preacher.name == new_name, Preacher.id != preacher.id).first()
        if existing:
            flash(f"Another preacher is already named '{new_name}'.")
            return redirect(url_for("admin.preacher_edit", preacher_id=preacher_id))

        preacher.name = new_name
        db.session.commit()
        flash(f"Preacher renamed to '{preacher.name}'.")
        return redirect(url_for("admin.preachers"))

    return render_template("admin/preacher_edit.html", preacher=preacher)


@bp.route("/preachers/<int:preacher_id>/delete", methods=["POST"])
@login_required
@admin_role_required
def preacher_delete(preacher_id):
    preacher = Preacher.query.get_or_404(preacher_id)
    has_donors = Donor.query.filter_by(connected_preacher_id=preacher.id).first() is not None
    if has_donors:
        flash(
            f"Can't delete '{preacher.name}' -- donors are still connected to them. "
            "Deactivate instead to hide them from the Connected Preacher dropdown, "
            "or reassign those donors first."
        )
        return redirect(url_for("admin.preachers"))

    db.session.delete(preacher)
    db.session.commit()
    flash(f"Preacher '{preacher.name}' deleted.")
    return redirect(url_for("admin.preachers"))


def _next_occurrence(date_value, today):
    """Days until the next anniversary of date_value's month/day, ignoring
    the year -- handles the year-boundary wraparound (e.g. today is Dec
    25, the event is Jan 3) by trying this year first and rolling over to
    next year if that's already passed. A Feb 29 birthday in a non-leap
    target year is treated as Feb 28 rather than raising."""
    for year in (today.year, today.year + 1):
        try:
            candidate = date_value.replace(year=year)
        except ValueError:
            candidate = date_value.replace(year=year, day=28)
        if candidate >= today:
            return (candidate - today).days, candidate
    return None, None  # unreachable -- the year+1 branch always satisfies candidate >= today


@bp.route("/donor-insights")
@login_required
def donor_insights():
    """Aggregate reports over the relationship-management fields on Donor
    (donor type, connected preacher, donation frequency, family DOBs /
    anniversary) -- totals by type and by preacher, an unassigned-donors
    count, a donation-frequency breakdown, and an upcoming birthdays/
    anniversaries list for relationship building.
    """
    # --- donors + donation totals by donor type ---
    donor_type_counts = {t: 0 for t in DONOR_TYPES}
    donor_type_counts["_uncategorised"] = 0
    for donor_type, count in db.session.query(Donor.donor_type, func.count(Donor.id)).group_by(Donor.donor_type).all():
        if donor_type in donor_type_counts:
            donor_type_counts[donor_type] = count
        else:
            donor_type_counts["_uncategorised"] += count

    donor_type_totals = {t: 0.0 for t in DONOR_TYPES}
    rows = (
        db.session.query(Donor.donor_type, func.coalesce(func.sum(Donation.amount), 0))
        .join(Donation, Donation.donor_id == Donor.id)
        .filter(Donation.status == "success")
        .group_by(Donor.donor_type)
        .all()
    )
    for donor_type, total in rows:
        if donor_type in donor_type_totals:
            donor_type_totals[donor_type] = float(total)

    # --- donors + donation totals by preacher ---
    preacher_rows = (
        db.session.query(
            Preacher.id, Preacher.name,
            func.count(func.distinct(Donor.id)),
            func.coalesce(func.sum(Donation.amount), 0),
        )
        .join(Donor, Donor.connected_preacher_id == Preacher.id)
        .outerjoin(Donation, (Donation.donor_id == Donor.id) & (Donation.status == "success"))
        .group_by(Preacher.id, Preacher.name)
        .order_by(Preacher.name)
        .all()
    )
    preacher_stats = [
        {"id": pid, "name": name, "donor_count": donor_count, "total": float(total)}
        for pid, name, donor_count, total in preacher_rows
    ]
    unassigned_count = Donor.query.filter(Donor.connected_preacher_id.is_(None)).count()

    # --- donation frequency breakdown ---
    frequency_counts = {f: 0 for f in DONATION_FREQUENCIES}
    for freq, count in db.session.query(Donor.donation_frequency, func.count(Donor.id)).group_by(Donor.donation_frequency).all():
        if freq in frequency_counts:
            frequency_counts[freq] = count

    # --- upcoming birthdays / anniversaries ---
    upcoming_days = request.args.get("days", 30, type=int)
    # now_ist(), not datetime.date.today() -- the server's own clock is UTC
    # (e.g. on Render), so "today" would read as the wrong calendar date for
    # the ~5.5 hours a day (roughly 12:00 AM-5:30 AM IST) where UTC hasn't
    # rolled over to the same day yet.
    today = now_ist().date()
    donors_with_dates = Donor.query.filter(
        db.or_(
            Donor.dob.isnot(None), Donor.father_dob.isnot(None), Donor.mother_dob.isnot(None),
            Donor.wife_dob.isnot(None), Donor.marriage_anniversary.isnot(None),
        )
    ).all()

    upcoming = []
    for donor in donors_with_dates:
        for label, date_value in [
            ("Donor's Birthday", donor.dob), ("Marriage Anniversary", donor.marriage_anniversary),
            ("Father's Birthday", donor.father_dob), ("Mother's Birthday", donor.mother_dob),
            ("Wife's Birthday", donor.wife_dob),
        ]:
            if not date_value:
                continue
            days_until, occurrence_date = _next_occurrence(date_value, today)
            if days_until is not None and days_until <= upcoming_days:
                upcoming.append({"donor": donor, "label": label, "days_until": days_until, "date": occurrence_date})
    upcoming.sort(key=lambda u: u["days_until"])

    return render_template(
        "admin/donor_insights.html",
        donor_type_counts=donor_type_counts, donor_type_totals=donor_type_totals,
        donor_type_labels=DONOR_TYPE_LABELS, donor_types=DONOR_TYPES,
        preacher_stats=preacher_stats, unassigned_count=unassigned_count,
        frequency_counts=frequency_counts, donation_frequency_labels=DONATION_FREQUENCY_LABELS,
        donation_frequencies=DONATION_FREQUENCIES,
        upcoming=upcoming, upcoming_days=upcoming_days,
    )


@bp.route("/birthdays")
@login_required
def birthdays():
    """Dedicated birthday calendar -- donor's own DOB only (family DOBs /
    marriage anniversaries get the fuller treatment on Donor Insights,
    which covers a wider set of relationship-building dates). Today's
    birthdays are broken out separately; everything else within the
    selected window is listed soonest-first."""
    # now_ist(), not datetime.date.today() -- the server's own clock is UTC
    # (e.g. on Render), so "today" would read as the wrong calendar date for
    # the ~5.5 hours a day (roughly 12:00 AM-5:30 AM IST) where UTC hasn't
    # rolled over to the same day yet.
    today = now_ist().date()
    window_days = request.args.get("days", 60, type=int)

    donors_with_dob = Donor.query.filter(Donor.dob.isnot(None)).all()

    all_birthdays = []
    for donor in donors_with_dob:
        days_until, occurrence_date = _next_occurrence(donor.dob, today)
        all_birthdays.append({"donor": donor, "days_until": days_until, "date": occurrence_date})
    all_birthdays.sort(key=lambda u: u["days_until"])

    birthdays_today = [u for u in all_birthdays if u["days_until"] == 0]
    birthdays_upcoming = [u for u in all_birthdays if 0 < u["days_until"] <= window_days]

    return render_template(
        "admin/birthdays.html",
        birthdays_today=birthdays_today,
        birthdays_upcoming=birthdays_upcoming,
        window_days=window_days,
        donors_with_dob_count=len(donors_with_dob),
        total_donors=Donor.query.count(),
    )


@bp.route("/settings/backup")
@login_required
@admin_role_required
def data_backup():
    """On-demand full data backup (donors/donations/lookup lists as CSV,
    zipped -- see backup_utils.build_backup_zip). Complements
    backup_data.py, which is meant to run automatically on a weekly
    schedule (Render Cron Job or any external scheduler) -- this page is
    for "I want one right now" without waiting for that."""
    backups_dir = current_app.config.get("BACKUP_DIR") or os.path.join(
        current_app.root_path, "instance", "backups"
    )
    recent_backups = []
    if os.path.isdir(backups_dir):
        recent_backups = sorted(
            (f for f in os.listdir(backups_dir) if f.startswith("temple_data_backup_") and f.endswith(".zip")),
            reverse=True,
        )[:10]
    return render_template(
        "admin/data_backup.html", recent_backups=recent_backups,
        backup_email=current_app.config.get("BACKUP_EMAIL") or current_app.config.get("ORG_CONTACT_EMAIL"),
        smtp_configured=bool(current_app.config.get("SMTP_HOST")),
    )


@bp.route("/settings/backup/download")
@login_required
@admin_role_required
def download_backup():
    filename, zip_bytes = build_backup_zip()
    return Response(
        zip_bytes, mimetype="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@bp.route("/settings/backup/run", methods=["POST"])
@login_required
@admin_role_required
def trigger_backup():
    """"Run Backup Now" button on the Data Backup page -- runs the exact
    same save-to-disk + prune + email routine as the weekly backup_data.py
    Cron Job (see backup_utils.run_backup), but on demand from this web
    service's own filesystem instead of waiting for Sunday 2 AM UTC."""
    result = run_backup(current_app, send_email=True)

    size_mb = result["size_bytes"] / (1024 * 1024)
    flash(f"Backup created: {result['filename']} ({size_mb:.2f} MB), saved to instance/backups/.")
    if result["pruned"]:
        flash(f"Pruned {len(result['pruned'])} old backup(s) beyond retention limit.")
    if result["email_sent"]:
        flash(f"Emailed backup to {result['emailed_to']}.")
    elif result["email_skipped_reason"]:
        flash(f"Backup not emailed: {result['email_skipped_reason']}.")

    return redirect(url_for("admin.data_backup"))


@bp.route("/export/10bd")
@login_required
def export_10bd():
    fy = request.args.get("fy") or get_financial_year()
    rows = (
        db.session.query(Donation, Donor, Campaign)
        .join(Donor, Donation.donor_id == Donor.id)
        .join(Campaign, Donation.campaign_id == Campaign.id)
        .filter(Donation.status == "success", Donation.financial_year == fy)
        .order_by(Donation.donation_date)
        .all()
    )
    # Filtered in Python rather than at the query level (Campaign.is_80g)
    # because 80G-eligibility is per-donation for Live To Give (donor picks
    # 80G/Non-80G per donation) -- see Donation.effective_is_80g.
    rows = [(donation, donor, campaign) for donation, donor, campaign in rows if donation.effective_is_80g]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Receipt No", "Date", "Donor Name", "Address", "City", "State", "Pincode",
        "PAN", "Amount", "Payment Mode", "Campaign",
    ])
    for donation, donor, campaign in rows:
        writer.writerow(csv_safe_row([
            donation.receipt_number,
            donation.donation_date.strftime("%d-%m-%Y"),
            donor.full_name,
            donor.address or "",
            donor.city or "",
            donor.state or "",
            donor.pincode or "",
            donor.pan or "",
            float(donation.amount),
            donation.payment_mode,
            campaign.name,
        ]))

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename=Form10BD_data_{fy}.csv"},
    )


@bp.route("/export/collections")
@login_required
def export_collections():
    fy = request.args.get("fy") or get_financial_year()
    rows = (
        db.session.query(Donation, Donor, Campaign)
        .join(Donor, Donation.donor_id == Donor.id)
        .join(Campaign, Donation.campaign_id == Campaign.id)
        .filter(Donation.status == "success", Donation.financial_year == fy)
        .order_by(Donation.donation_date)
        .all()
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Receipt No", "Date", "Donor Name", "Phone", "WhatsApp", "Email", "PAN", "Amount",
        "Payment Mode", "Campaign", "80G Eligible",
    ])
    for donation, donor, campaign in rows:
        writer.writerow(csv_safe_row([
            donation.receipt_number,
            donation.donation_date.strftime("%d-%m-%Y"),
            donor.full_name,
            donor.phone or "",
            donor.whatsapp_number or "",
            donor.email or "",
            donor.pan or "",
            float(donation.amount),
            donation.payment_mode,
            campaign.name,
            "Yes" if donation.effective_is_80g else "No",
        ]))

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename=Collections_{fy}.csv"},
    )


@bp.route("/export/donations")
@login_required
def export_donations():
    """CSV export of the Donations Log, honoring whatever campaign/mode/
    status filters are currently applied on that page -- unlike the
    on-screen table (paginated to DONATIONS_PER_PAGE rows), this exports
    every matching row. Carries the same full detail set as the Donations
    Log's "Full details" modal (donor contact info, PAN, address, payment
    reference, campaign-specific purpose, 80G status, who recorded it, and
    cancellation info where relevant) so this file alone answers "what did
    this donor actually enter" without having to click into each row.
    """
    query, filters = _apply_donations_filters(Donation.query)
    status = filters["status"]
    query, _sort = _sorted_donations_query(query, request.args.get("sort", "-date"))

    rows = query.all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Receipt No", "Date", "Status", "Donor Name", "Phone", "WhatsApp", "Email", "PAN",
        "Address", "City", "State", "Pincode",
        "Amount", "Payment Mode", "Reference", "Order ID", "Campaign", "Specific Purpose", "80G Eligible",
        "Camp", "Batch",
        "Recorded By", "Remarks", "Cancelled At", "Cancelled By", "Cancellation Reason",
    ])
    for d in rows:
        donor = d.donor
        specific_purpose = d.specific_purpose
        # Online donations carry a real time-of-day (set the instant the
        # payment was confirmed); offline entries are always saved at
        # midnight since only a date is captured for those, so a time
        # component would be misleading noise -- see the same convention
        # in the Donations Log table and detail modal.
        date_str = (
            to_ist(d.donation_date).strftime("%d-%m-%Y %H:%M")
            if d.payment_mode == "online"
            else d.donation_date.strftime("%d-%m-%Y")
        )
        writer.writerow(csv_safe_row([
            d.receipt_number or "",
            date_str,
            d.status,
            donor.full_name,
            donor.phone or "",
            donor.whatsapp_number or "",
            donor.email or "",
            donor.pan or "",
            donor.address or "",
            donor.city or "",
            donor.state or "",
            donor.pincode or "",
            float(d.amount),
            d.payment_mode,
            d.reference_display or "",
            d.razorpay_order_id or "",
            d.campaign.name,
            specific_purpose,
            "Yes" if d.effective_is_80g else "No",
            d.camp_name or "",
            d.batch_name or "",
            d.recorded_by or "",
            d.remarks or "",
            to_ist(d.cancelled_at).strftime("%d-%m-%Y %H:%M") if d.cancelled_at else "",
            d.cancelled_by or "",
            d.cancellation_reason or "",
        ]))

    filename_status = status if status != "all" else "all-statuses"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename=Donations_{filename_status}.csv"},
    )


@bp.route("/export/monthly")
@login_required
def export_monthly_report():
    """Full monthly donor report -- every successful donation in the given
    calendar month, with the donor's complete details (contact, address,
    PAN) alongside the donation's own details (amount, mode, reference,
    campaign, receipt number, 80G eligibility). Broader than
    export_collections (which is FY-wide and has fewer donor fields) --
    this is meant to be handed off as one month's complete donor list,
    e.g. for a mailing list or an external audit request.

    Cancelled donations are excluded for free by the existing
    status == "success" filter (see Donation.status / the cancellation
    system), same as every other export/total in this file.
    """
    # now_ist(), not datetime.date.today() -- the server's own clock is UTC
    # (e.g. on Render), so "today" would read as the wrong calendar date for
    # the ~5.5 hours a day (roughly 12:00 AM-5:30 AM IST) where UTC hasn't
    # rolled over to the same day yet.
    today = now_ist().date()
    year = request.args.get("year", today.year, type=int)
    month = request.args.get("month", today.month, type=int)
    if not (1 <= month <= 12):
        month = today.month

    start = datetime.date(year, month, 1)
    end = datetime.date(year + (1 if month == 12 else 0), 1 if month == 12 else month + 1, 1)

    rows = (
        db.session.query(Donation, Donor, Campaign)
        .join(Donor, Donation.donor_id == Donor.id)
        .join(Campaign, Donation.campaign_id == Campaign.id)
        .filter(
            Donation.status == "success",
            Donation.donation_date >= start,
            Donation.donation_date < end,
        )
        .order_by(Donation.donation_date)
        .all()
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Receipt No", "Date", "Donor Name", "Phone", "WhatsApp", "Email", "PAN",
        "Address", "City", "State", "Pincode",
        "Amount", "Payment Mode", "Reference", "Campaign", "80G Eligible", "Recorded By",
    ])
    for donation, donor, campaign in rows:
        writer.writerow(csv_safe_row([
            donation.receipt_number,
            donation.donation_date.strftime("%d-%m-%Y"),
            donor.full_name,
            donor.phone or "",
            donor.whatsapp_number or "",
            donor.email or "",
            donor.pan or "",
            donor.address or "",
            donor.city or "",
            donor.state or "",
            donor.pincode or "",
            float(donation.amount),
            donation.payment_mode,
            donation.reference_display or "",
            campaign.name,
            "Yes" if donation.effective_is_80g else "No",
            donation.recorded_by or "",
        ]))

    month_label = start.strftime("%Y-%m")
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename=Monthly_Donor_Report_{month_label}.csv"},
    )


@bp.route("/donors/lapsed")
@login_required
def lapsed_donors():
    """Donors who gave in each of the last 3 months before last month, but
    NOT last month -- i.e. likely-recurring donors who may need a follow-up."""
    # now_ist(), not datetime.date.today() -- the server's own clock is UTC
    # (e.g. on Render), so "today" would read as the wrong calendar date for
    # the ~5.5 hours a day (roughly 12:00 AM-5:30 AM IST) where UTC hasn't
    # rolled over to the same day yet.
    today = now_ist().date()

    def month_key(d):
        return (d.year, d.month)

    def months_ago(base, i):
        total_months = base.year * 12 + (base.month - 1) - i
        y, m = divmod(total_months, 12)
        return datetime.date(y, m + 1, 1)

    start_of_this_month = today.replace(day=1)
    months_back = [month_key(months_ago(start_of_this_month, i)) for i in range(1, 5)]
    last_month_key = months_back[0]

    donors_all = Donor.query.all()
    lapsed = []
    for donor in donors_all:
        months_donated = set()
        for d in donor.donations.filter_by(status="success"):
            months_donated.add(month_key(d.donation_date.date()))
        prior_three = set(months_back[1:4])
        if prior_three.issubset(months_donated) and last_month_key not in months_donated:
            lapsed.append(donor)

    return render_template("admin/lapsed_donors.html", donors=lapsed)


# ---------------------------------------------------------------------------
# IYF Camps
# ---------------------------------------------------------------------------
# Donations collected from students at IYF camps. Structurally these are
# ordinary offline donations -- they go through the same
# _create_offline_donation() path, get real receipt numbers from the same
# counter, and show up in the Donations Log like everything else. The only
# thing that makes them camp donations is the camp_name/batch_name pair on
# the row, which is what the per-camp totals group by.
#
# Two deliberate differences from the Offline Donation tab:
#   - They're always filed against the "IYF Camps" campaign, created on
#     first use. Camp money is its own stream; putting it under an existing
#     campaign would corrupt that campaign's totals.
#   - Nothing is emailed or WhatsApped, ever. Most of this data arrives as a
#     Zoho export of payments students made weeks earlier, and messaging
#     hundreds of them retrospectively would be worse than useless. Receipts
#     are still generated and downloadable from the admin side.

IYF_CAMP_IMPORT_REQUIRED_COLUMNS = ["full_name", "amount", "camp_name"]
IYF_CAMP_IMPORT_COLUMNS = [
    "full_name", "amount", "camp_name", "batch_name", "donation_date",
    "payment_mode", "phone", "email", "bank_transaction_id", "remarks",
]

# Shared by the CSV template and the .xlsx one, so the two can't drift
# into showing different examples of the same thing.
IYF_CAMP_DEMO_ROWS = [
    {
        "full_name": "Ravi Sharma", "amount": "1100", "camp_name": "Utkarsha 2026",
        "batch_name": "Batch A", "donation_date": "2026-08-01", "payment_mode": "cash",
        "phone": "9876543210", "email": "ravi@example.com",
        "bank_transaction_id": "", "remarks": "",
    },
    {
        "full_name": "Anita Verma", "amount": "2100", "camp_name": "Utkarsha 2026",
        "batch_name": "Batch B", "donation_date": "2026-08-02", "payment_mode": "online",
        "phone": "9812345678", "email": "",
        "bank_transaction_id": "pay_TO5ASGCNZOi4fP", "remarks": "Paid via Zoho",
    },
]


@bp.route("/iyf-camps")
@login_required
def iyf_camps():
    camps = Camp.query.order_by(Camp.is_active.desc(), Camp.name).all()
    batch_names = _known_batch_names()

    recent = (
        Donation.query.filter(Donation.camp_name.isnot(None))
        .order_by(Donation.id.desc())
        .limit(25)
        .all()
    )

    return render_template(
        "admin/iyf_camps.html",
        camps=camps, batch_names=batch_names, recent=recent,
        today=datetime.date.today(),
        import_columns=IYF_CAMP_IMPORT_COLUMNS,
        required_columns=IYF_CAMP_IMPORT_REQUIRED_COLUMNS,
        tab=request.args.get("tab", "single"),
    )


@bp.route("/iyf-camps/single", methods=["POST"])
@login_required
def iyf_camp_single():
    form = request.form

    # Resolved against the Camp list rather than taken as typed, so the
    # stored name is always the camp's own spelling.
    camp_name = _resolve_camp_name(form.get("camp_name"))
    if not camp_name:
        flash("Please choose a camp from the list. Add it under Manage Camps first if it's missing.")
        return redirect(url_for("admin.iyf_camps"))

    try:
        amount = float(form.get("amount") or 0)
    except ValueError:
        amount = 0
    if amount <= 0:
        flash("Enter a valid amount.")
        return redirect(url_for("admin.iyf_camps"))

    if not (form.get("full_name") or "").strip():
        flash("Student name is required.")
        return redirect(url_for("admin.iyf_camps"))

    # Phone is optional here (unlike the public forms): a camp register
    # often has only a name against a cash payment. But a value that *is*
    # given still has to be a real number, or donor dedup starts matching
    # unrelated students to each other on a mistyped phone.
    phone = (form.get("phone") or "").strip()
    if phone and not is_valid_phone(phone):
        flash("That phone number doesn't look right. Please enter a 10-digit mobile number, or leave it blank.")
        return redirect(url_for("admin.iyf_camps"))

    donation_date_str = form.get("donation_date")
    try:
        donation_date = (
            datetime.datetime.strptime(donation_date_str, "%Y-%m-%d")
            if donation_date_str else datetime.datetime.utcnow()
        )
    except ValueError:
        flash("That donation date doesn't look right.")
        return redirect(url_for("admin.iyf_camps"))

    result = _create_offline_donation(
        donor_data=form,
        campaign=_iyf_camp_campaign(),
        amount=amount,
        payment_mode=form.get("payment_mode", "cash"),
        donation_date=donation_date,
        recorded_by=current_user.username,
        bank_transaction_id=form.get("bank_transaction_id"),
        remarks=form.get("remarks"),
        camp_name=camp_name,
        batch_name=_normalize_camp_text(form.get("batch_name")),
        # See the module comment above -- camp entries never notify.
        send_notifications=False,
    )

    if not result["ok"]:
        flash(result["error"])
        return redirect(url_for("admin.iyf_camps"))

    log_activity(
        "iyf_camp_donation", target_type="donation", target_id=result["donation"].id,
        details=f"Camp '{camp_name}' donation Rs. {amount:.2f}, receipt {result['receipt_number']}",
    )
    db.session.commit()

    msg = f"Recorded Rs. {amount:.0f} for {camp_name} -- receipt {result['receipt_number']}."
    if not result["pdf_ok"]:
        msg += " (The receipt PDF couldn't be generated; it can be downloaded again later.)"
    flash(msg)
    return redirect(url_for("admin.iyf_camps"))


@bp.route("/iyf-camps/bulk", methods=["POST"])
@login_required
def iyf_camp_bulk():
    file = request.files.get("csv_file")
    if not file or not file.filename:
        flash("Please choose a CSV file to upload.")
        return redirect(url_for("admin.iyf_camps", tab="bulk"))

    try:
        reader = _table_from_upload(file)
        fieldnames = {(f or "").strip() for f in (reader.fieldnames or [])}
    except _UploadReadError as exc:
        # Says what's actually wrong with the file. The generic message
        # below used to cover this case too, which told someone whose
        # Excel file needed re-saving to upload a CSV instead.
        flash(str(exc))
        return redirect(url_for("admin.iyf_camps", tab="bulk"))
    except Exception:
        flash("Couldn't read that file -- please upload a CSV or Excel (.xlsx) file.")
        return redirect(url_for("admin.iyf_camps", tab="bulk"))

    missing = [c for c in IYF_CAMP_IMPORT_REQUIRED_COLUMNS if c not in fieldnames]
    if missing:
        flash(
            "That CSV is missing required column(s): " + ", ".join(missing)
            + ". Download the template below for the full column list."
        )
        return redirect(url_for("admin.iyf_camps", tab="bulk"))

    preview = request.form.get("action") == "preview"  # dry run -- see bulk_import_donations
    # create=False on a dry run: this call commits, so previewing used to
    # create the IYF Camps campaign while reporting that nothing had been
    # saved. May be None here; only the non-preview path dereferences it.
    campaign = _iyf_camp_campaign(create=not preview)
    results, created = [], 0
    ambiguous_dates = []  # see _parse_import_date

    for line_num, row in enumerate(reader, start=2):
        row = {(k or "").strip(): (v or "").strip() for k, v in row.items() if k}
        row_errors = []

        full_name = row.get("full_name", "")
        if not full_name:
            row_errors.append("full_name is required")

        raw_camp = _normalize_camp_text(row.get("camp_name"))
        camp_name = _resolve_camp_name(raw_camp)
        if not raw_camp:
            row_errors.append("camp_name is required")
        elif not camp_name:
            # Named rather than silently created: an unrecognised camp in
            # an export is nearly always a spelling difference, and
            # inventing a camp for it would split that camp's total in two
            # -- the exact failure the managed list exists to prevent.
            row_errors.append(
                f"no camp named '{raw_camp}' -- add it under Manage Camps, or correct the spelling"
            )

        try:
            amount = float(row.get("amount") or 0)
            if amount <= 0:
                row_errors.append("amount must be greater than 0")
        except ValueError:
            amount = 0
            row_errors.append(f"invalid amount '{row.get('amount')}'")

        phone = row.get("phone", "")
        if phone and not is_valid_phone(phone):
            # Not fatal on purpose: a Zoho export can carry a phone mangled
            # into scientific notation by a spreadsheet round-trip, and
            # losing the whole donation over an unusable phone number would
            # be the wrong trade. Dropped with a note instead, so the money
            # is still recorded against the right student and camp.
            row_errors.append(f"phone '{phone}' isn't a valid 10-digit mobile -- imported without it")
            phone = ""

        payment_mode = (row.get("payment_mode") or "cash").lower()
        if payment_mode not in ("cash", "cheque", "bank_transfer", "online"):
            row_errors.append(f"payment_mode must be cash, cheque, bank_transfer, or online (got '{payment_mode}')")

        donation_date = _parse_import_date(
            row.get("donation_date"), "donation_date", row_errors, ambiguous=ambiguous_dates)

        # Mirrors the check inside _create_offline_donation so a preview
        # can't promise a row the real run will refuse.
        reference_error = _payment_reference_error(
            payment_mode, None, row.get("bank_transaction_id"))
        if reference_error:
            row_errors.append(reference_error)

        fatal = [e for e in row_errors if "imported without it" not in e]
        if fatal:
            results.append({"line": line_num, "name": full_name or "(blank)", "ok": False, "errors": fatal})
            continue

        if preview:
            results.append({
                "line": line_num, "name": full_name, "ok": True, "preview": True,
                # The constant, not campaign.name -- campaign is None on a
                # dry run when the IYF Camps campaign doesn't exist yet.
                "amount": amount, "campaign": IYF_CAMP_CAMPAIGN_NAME,
                "donation_date": (
                    datetime.datetime.combine(donation_date, datetime.time())
                    if donation_date else None
                ),
                "payment_mode": payment_mode,
                "warnings": [e for e in row_errors if "imported without it" in e],
            })
            created += 1
            continue

        result = _create_offline_donation(
            donor_data={
                "full_name": full_name, "phone": phone, "email": row.get("email", ""),
            },
            campaign=campaign,
            amount=amount,
            payment_mode=payment_mode,
            donation_date=(
                datetime.datetime.combine(donation_date, datetime.time())
                if donation_date else datetime.datetime.utcnow()
            ),
            recorded_by=current_user.username,
            bank_transaction_id=row.get("bank_transaction_id"),
            remarks=row.get("remarks"),
            camp_name=camp_name,
            batch_name=_normalize_camp_text(row.get("batch_name")),
            send_notifications=False,
        )

        if not result["ok"]:
            results.append({"line": line_num, "name": full_name, "ok": False, "errors": [result["error"]]})
            continue

        created += 1
        results.append({
            "line": line_num, "name": full_name, "ok": True,
            "receipt_number": result["receipt_number"],
            "warnings": [e for e in row_errors if "imported without it" in e],
        })

    if preview:
        flash(
            f"Preview only -- nothing has been saved. {created} row(s) would be "
            f"imported, {len(results) - created} skipped."
            + _ambiguous_date_warning(ambiguous_dates)
        )
    else:
        log_activity("iyf_camp_bulk_import", details=f"Imported {created} camp donation(s) from CSV")
        db.session.commit()

    camps = Camp.query.order_by(Camp.is_active.desc(), Camp.name).all()
    batch_names = _known_batch_names()
    return render_template(
        "admin/iyf_camps.html",
        camps=camps, batch_names=batch_names,
        recent=Donation.query.filter(Donation.camp_name.isnot(None))
                             .order_by(Donation.id.desc()).limit(25).all(),
        today=datetime.date.today(),
        import_columns=IYF_CAMP_IMPORT_COLUMNS,
        required_columns=IYF_CAMP_IMPORT_REQUIRED_COLUMNS,
        tab="bulk", bulk_preview=preview, bulk_summary=_preview_summary(results),
        bulk_results=results, bulk_created=created,
        bulk_skipped=len(results) - created,
    )


@bp.route("/iyf-camps/template.csv")
@login_required
def iyf_camp_template():
    """A CSV with the exact headers the importer expects, plus two example
    rows -- easier to hand someone than a column list they have to
    transcribe, and it round-trips through Excel without surprises."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=IYF_CAMP_IMPORT_COLUMNS)
    writer.writeheader()
    writer.writerows(IYF_CAMP_DEMO_ROWS)
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=iyf_camp_import_template.csv"},
    )


@bp.route("/iyf-camps/template.xlsx")
@login_required
def iyf_camp_template_xlsx():
    return _demo_workbook(
        IYF_CAMP_IMPORT_COLUMNS, IYF_CAMP_DEMO_ROWS, "iyf_camp_import_template.xlsx")


# --- Camp list management -------------------------------------------------
# The camps offered in the entry dropdown. Kept deliberately separate from
# the donations themselves: Donation.camp_name is a copy of the name taken
# when the donation was recorded, so deleting a camp here never disturbs
# what it collected. See the Camp model docstring for why.

@bp.route("/iyf-camps/manage", methods=["GET", "POST"])
@login_required
@admin_role_required
def camps():
    if request.method == "POST":
        name = _normalize_camp_text(request.form.get("name"))
        if not name:
            flash("Camp name is required.")
            return redirect(url_for("admin.camps"))

        # Case-insensitive, because "utkarsha 2026" and "Utkarsha 2026"
        # being two camps is exactly what this list exists to prevent.
        if Camp.query.filter(func.lower(Camp.name) == name.lower()).first():
            flash(f"There's already a camp called '{name}'.")
            return redirect(url_for("admin.camps"))

        camp = Camp(name=name)
        db.session.add(camp)
        db.session.commit()
        log_activity("camp_create", target_type="camp", target_id=camp.id, details=f"Created camp '{name}'")
        db.session.commit()
        flash(f"Camp '{name}' added.")
        return redirect(url_for("admin.camps"))

    camps_list = Camp.query.order_by(Camp.is_active.desc(), Camp.name).all()

    # How much each camp has collected, so staff can see what they'd be
    # detaching before deleting one.
    totals = dict(
        db.session.query(Donation.camp_name, func.coalesce(func.sum(Donation.amount), 0))
        .filter(Donation.camp_name.isnot(None), Donation.status == "success")
        .group_by(Donation.camp_name).all()
    )
    counts = dict(
        db.session.query(Donation.camp_name, func.count(Donation.id))
        .filter(Donation.camp_name.isnot(None), Donation.status == "success")
        .group_by(Donation.camp_name).all()
    )
    return render_template("admin/camps.html", camps=camps_list, totals=totals, counts=counts)


@bp.route("/iyf-camps/manage/<int:camp_id>/edit", methods=["POST"])
@login_required
@admin_role_required
def camp_edit(camp_id):
    camp = Camp.query.get_or_404(camp_id)
    new_name = _normalize_camp_text(request.form.get("name"))
    if not new_name:
        flash("Camp name is required.")
        return redirect(url_for("admin.camps"))

    clash = Camp.query.filter(func.lower(Camp.name) == new_name.lower(), Camp.id != camp.id).first()
    if clash:
        flash(f"There's already a camp called '{clash.name}'.")
        return redirect(url_for("admin.camps"))

    old_name = camp.name
    camp.name = new_name
    camp.is_active = request.form.get("is_active") == "yes"

    # Carry the rename onto donations already recorded under the old name.
    # Without this a corrected spelling would split one camp's history into
    # two totals -- the old name still sitting on past donations, the new
    # one collecting everything from here. This is the whole reason renames
    # go through a form rather than being edited row by row.
    renamed = 0
    if old_name != new_name:
        renamed = (
            Donation.query.filter(Donation.camp_name == old_name)
            .update({Donation.camp_name: new_name}, synchronize_session=False)
        )

    db.session.commit()
    log_activity(
        "camp_edit", target_type="camp", target_id=camp.id,
        details=f"Renamed camp '{old_name}' to '{new_name}'" if old_name != new_name
                else f"Updated camp '{new_name}'",
    )
    db.session.commit()

    msg = f"Camp updated to '{new_name}'."
    if renamed:
        msg += f" {renamed} existing donation(s) moved with it."
    flash(msg)
    return redirect(url_for("admin.camps"))


@bp.route("/iyf-camps/manage/<int:camp_id>/delete", methods=["POST"])
@login_required
@admin_role_required
def camp_delete(camp_id):
    camp = Camp.query.get_or_404(camp_id)
    name = camp.name

    # Safe to delete outright whatever has been collected: the donations
    # carry the camp name themselves, so they keep reporting under it. This
    # only removes the camp from the entry dropdown.
    kept = Donation.query.filter(Donation.camp_name == name).count()

    db.session.delete(camp)
    db.session.commit()
    log_activity(
        "camp_delete", target_type="camp", target_id=camp_id,
        details=f"Deleted camp '{name}' ({kept} donation(s) retained under that name)",
    )
    db.session.commit()

    msg = f"Camp '{name}' removed from the list."
    if kept:
        msg += f" Its {kept} donation(s) and their totals are unaffected."
    flash(msg)
    return redirect(url_for("admin.camps"))


# --- Camp collections report ----------------------------------------------

def _camp_report_filters():
    """Resolve the camp/date filters shared by the collections report and
    its exports, so an export always matches what's on screen.

    Returns (query, filters). Only successful donations count -- cancelled
    ones are money that came back out again, and a report used to reconcile
    what a camp actually raised must not include them.
    """
    query = Donation.query.filter(
        Donation.camp_name.isnot(None), Donation.status == "success"
    )

    camp = (request.args.get("camp") or "").strip()
    if camp:
        query = query.filter(Donation.camp_name == camp)

    date_from_raw = (request.args.get("date_from") or "").strip()
    date_to_raw = (request.args.get("date_to") or "").strip()
    try:
        if date_from_raw:
            query = query.filter(
                Donation.donation_date >= datetime.datetime.strptime(date_from_raw, "%Y-%m-%d")
            )
        if date_to_raw:
            # donation_date carries a time for online payments, so a plain
            # <= against midnight would drop everything later that day.
            query = query.filter(
                Donation.donation_date
                < datetime.datetime.strptime(date_to_raw, "%Y-%m-%d") + datetime.timedelta(days=1)
            )
    except ValueError:
        date_from_raw = date_to_raw = ""

    return query, {"camp": camp, "date_from": date_from_raw, "date_to": date_to_raw}


def _camp_totals(query):
    return (
        query.with_entities(
            Donation.camp_name,
            func.count(Donation.id),
            func.coalesce(func.sum(Donation.amount), 0),
        )
        .group_by(Donation.camp_name)
        .order_by(func.coalesce(func.sum(Donation.amount), 0).desc())
        .all()
    )


def _camp_monthly_totals(query):
    """Month-by-month totals.

    Grouped with extract() rather than strftime/date_trunc because those are
    dialect-specific -- the tests run on SQLite and production is Postgres,
    and a report that only works on one of them isn't a report.
    """
    rows = (
        query.with_entities(
            extract("year", Donation.donation_date).label("y"),
            extract("month", Donation.donation_date).label("m"),
            func.count(Donation.id),
            func.coalesce(func.sum(Donation.amount), 0),
        )
        .group_by("y", "m")
        .order_by("y", "m")
        .all()
    )
    out = []
    for y, m, count, total in rows:
        label = datetime.date(int(y), int(m), 1).strftime("%b %Y")
        out.append((label, int(count), float(total)))
    return out


@bp.route("/iyf-camps/collections")
@login_required
def iyf_camp_collections():
    query, filters = _camp_report_filters()
    return render_template(
        "admin/iyf_camp_collections.html",
        camp_totals=_camp_totals(query),
        monthly_totals=_camp_monthly_totals(query),
        camp_names=sorted(
            c for (c,) in db.session.query(Donation.camp_name)
            .filter(Donation.camp_name.isnot(None)).distinct().all() if c
        ),
        **filters,
    )


@bp.route("/iyf-camps/export/<report>.csv")
@login_required
def iyf_camp_export(report):
    """CSV of the collections report, honouring the on-screen filters.

    Three shapes, because they answer different questions: camp-wise for
    "what did each camp raise", monthly for "what came in when", and detail
    for the row-level data staff reconcile against a Zoho export or a
    cash book.
    """
    query, filters = _camp_report_filters()
    out = io.StringIO()
    writer = csv.writer(out)

    if report == "camp":
        writer.writerow(["Camp", "Donations", "Total"])
        for camp, count, total in _camp_totals(query):
            writer.writerow(csv_safe_row([camp, count, float(total)]))

    elif report == "monthly":
        writer.writerow(["Month", "Donations", "Total"])
        for label, count, total in _camp_monthly_totals(query):
            writer.writerow(csv_safe_row([label, count, total]))

    elif report == "detail":
        writer.writerow([
            "Date", "Receipt No", "Student", "Phone", "Email",
            "Camp", "Batch", "Amount", "Payment Mode", "Reference", "Recorded By",
        ])
        for d in query.order_by(Donation.camp_name, Donation.donation_date).all():
            donor = d.donor
            writer.writerow(csv_safe_row([
                d.donation_date.strftime("%d-%m-%Y") if d.donation_date else "",
                d.receipt_number or "",
                donor.full_name if donor else "",
                (donor.phone or "") if donor else "",
                (donor.email or "") if donor else "",
                d.camp_name or "",
                d.batch_name or "",
                float(d.amount),
                d.payment_mode or "",
                d.reference_display or "",
                d.recorded_by or "",
            ]))
    else:
        abort(404)

    # Name the file after what's in it -- a folder of "export.csv" is
    # useless a month later when someone asks which camp a figure came from.
    parts = ["IYF_Camps", report]
    if filters["camp"]:
        parts.append(re.sub(r"[^A-Za-z0-9]+", "_", filters["camp"]).strip("_"))
    if filters["date_from"] or filters["date_to"]:
        parts.append(f"{filters['date_from'] or 'start'}_to_{filters['date_to'] or 'today'}")
    filename = "-".join(parts) + ".csv"

    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename={filename}"},
    )


@bp.route("/settings/backup/restore", methods=["POST"])
@login_required
@admin_role_required
def restore_backup_upload():
    """Restore a backup ZIP from the Data Backup page.

    Restoring is the most destructive thing this admin panel can do, so
    three guards sit in front of it, all of which have to be got past
    deliberately:

      1. Preview is the default. The form's Preview button reports what
         would change and writes nothing. Applying takes a separate,
         differently-labelled button.
      2. Applying requires typing RESTORE. A misclick can't do it.
      3. A backup of the CURRENT database is taken and saved before any
         write. The CLI script tells you to do this by hand and trusts you
         to remember; here it's automatic, because the moment you need it
         is exactly the moment you'll have forgotten.

    Admin-only (staff can't reach it), on top of all of the above.
    """
    file = request.files.get("backup_zip")
    if not file or not file.filename:
        flash("Please choose a backup ZIP file.")
        return redirect(url_for("admin.data_backup"))

    apply_changes = request.form.get("mode") == "apply"
    wipe = request.form.get("wipe") == "yes"

    if apply_changes and (request.form.get("confirm") or "").strip().upper() != "RESTORE":
        flash("Type RESTORE in the confirmation box to apply a restore. Nothing was changed.")
        return redirect(url_for("admin.data_backup"))

    zip_bytes = file.read()

    safety_backup = None
    if apply_changes:
        # Before touching anything. If this fails, the restore doesn't run:
        # going ahead without a way back is not a trade worth making.
        try:
            safety_backup = run_backup(current_app, send_email=False)
        except Exception:
            current_app.logger.exception("Pre-restore safety backup failed")
            flash(
                "Couldn't take a safety backup of the current data, so the restore was "
                "cancelled. Nothing was changed."
            )
            return redirect(url_for("admin.data_backup"))

    try:
        result = restore_backup_zip(db, zip_bytes, wipe=wipe, dry_run=not apply_changes)
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Backup restore failed")
        flash(f"Restore failed: {exc}. Nothing was changed.")
        return redirect(url_for("admin.data_backup"))

    if not apply_changes:
        db.session.rollback()
    else:
        db.session.commit()
        log_activity(
            "backup_restore",
            details=f"Restored {file.filename}"
                    + (" (wipe first)" if wipe else " (upsert)")
                    + (f"; safety backup {safety_backup['filename']}" if safety_backup else ""),
        )
        db.session.commit()

    summary = ", ".join(
        f"{name.replace('.csv','')}: {ins} new / {upd} updated"
        for name, ins, upd in result["tables"]
    )
    if apply_changes:
        flash(f"Restore complete. {summary}.")
        if safety_backup:
            flash(f"A backup of the previous data was saved first as {safety_backup['filename']}.")
    else:
        flash(f"Preview only -- nothing was changed. This backup would apply: {summary}.")
    if result["missing"]:
        flash("Not in this backup (left as-is): " + ", ".join(result["missing"]) + ".")

    return redirect(url_for("admin.data_backup"))
